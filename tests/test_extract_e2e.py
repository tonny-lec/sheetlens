import json
import shutil
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import pytest
from typer.testing import CliRunner

from sheetlens.cli import app
from sheetlens.model import ir
import sheetlens.pipeline as pipeline
from sheetlens.pipeline import analyze
from sheetlens.question_ids import QuestionIdCatalog, resolve_answered_ids
from sheetlens.renderers.markdown import render_questions_md

runner = CliRunner()


def _build(wb):
    ws = wb.active
    ws.title = "見積入力"
    ws["A1"] = "見積書"
    ws.merge_cells("A1:C1")
    ws["A3"] = "顧客名"
    for r in range(11, 14):
        ws[f"C{r}"] = 2
        ws[f"D{r}"] = f"=VLOOKUP(A{r},単価マスタ!A:C,3,0)"
        ws[f"E{r}"] = f"=C{r}*D{r}"
    dv = DataValidation(type="list", formula1='"通常,特急"')
    dv.add("C5")
    ws.add_data_validation(dv)
    ws.conditional_formatting.add(
        "C11:C13",
        CellIsRule(
            operator="between",
            formula=["1", "10"],
            fill=PatternFill(fill_type="solid", fgColor="FFFF00"),
        ),
    )
    ws.conditional_formatting.add(
        "E11:E13",
        ColorScaleRule(
            start_type="min",
            start_color="FFFF0000",
            end_type="max",
            end_color="FF00FF00",
        ),
    )
    master = wb.create_sheet("単価マスタ")
    master["A1"] = "品名"


def _make_legacy_project(proj):
    raw = ir.Workbook.model_validate_json(
        (proj / "structure" / "raw.json").read_text(encoding="utf-8")
    )
    analysis = analyze(raw)
    (proj / "questions.md").write_text(
        render_questions_md(analysis.question_set.legacy_questions, set()),
        encoding="utf-8",
    )
    (proj / "question-ids.json").unlink()


def _write_legacy_annotation(proj):
    path = proj / "annotations" / "見積入力.yaml"
    path.write_text(
        "sheet: 見積入力\nquestions_answered: [q-001]\n",
        encoding="utf-8",
    )
    return path, path.read_bytes()


def _insert_sheet_first(src, title):
    wb = openpyxl.load_workbook(src)
    wb.create_sheet(title, 0)
    wb.save(src)


def _structure_bytes(proj):
    structure = proj / "structure"
    return {
        path.relative_to(structure).as_posix(): path.read_bytes()
        for path in structure.rglob("*")
        if path.is_file()
    }


def _project_bytes(proj):
    return {
        path.relative_to(proj).as_posix(): path.read_bytes()
        for path in proj.rglob("*")
        if path.is_file() and not path.is_symlink()
    }


def test_extract_generates_project(make_xlsx):
    src = make_xlsx(_build, name="見積管理.xlsx")
    result = runner.invoke(app, ["extract", str(src)])
    assert result.exit_code == 0, result.output
    proj = src.parent / "見積管理.sheetlens"
    for rel in (
        "manifest.json",
        "question-ids.json",
        "questions.md",
        "README.md",
        "structure/raw.json",
        "structure/sheet-見積入力.md",
        "annotations",
    ):
        assert (proj / rel).exists(), rel
    raw = json.loads((proj / "structure/raw.json").read_text(encoding="utf-8"))
    conditional_formats = raw["sheets"][0]["conditional_formats"]
    between = conditional_formats[0]
    assert between["formulas"] == ["1", "10"]
    assert "formula" not in between
    assert between["dxf"] is not None
    assert between["dxf"]["tag"].endswith("dxf")
    assert any(child["tag"].endswith("fill") for child in between["dxf"]["children"])
    color_scale = conditional_formats[1]
    assert color_scale["color_scale"] == {
        "conditions": [
            {"type": "min", "value": None, "gte": None},
            {"type": "max", "value": None, "gte": None},
        ],
        "colors": [
            {"type": "rgb", "value": "FFFF0000", "tint": 0.0},
            {"type": "rgb", "value": "FF00FF00", "tint": 0.0},
        ],
    }
    assert "formula" not in color_scale
    catalog = json.loads((proj / "question-ids.json").read_text(encoding="utf-8"))
    assert catalog["schema_version"] == 1
    assert catalog["generator_version"] == 2
    assert catalog["source_sha256"] == raw["sha256"]
    manifest = json.loads((proj / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["dependencies"]["見積入力"] == ["単価マスタ"]
    md = (proj / "structure/sheet-見積入力.md").read_text(encoding="utf-8")
    assert "[A1:C1 結合]" in md
    assert "=R[0]C[-2]*R[0]C[-1]" in md
    assert "通常" in md
    questions = (proj / "questions.md").read_text(encoding="utf-8")
    assert "dropdown_semantics" in questions and "sheet_role" in questions


def test_extract_preserves_annotations(make_xlsx):
    src = make_xlsx(_build, name="a.xlsx")
    proj = src.parent / "a.sheetlens"
    (proj / "annotations").mkdir(parents=True)
    keep = proj / "annotations" / "見積入力.yaml"
    malformed = b"sheet: [\xff\x00unterminated"
    keep.write_bytes(malformed)
    result = runner.invoke(app, ["extract", str(src)])
    assert result.exit_code == 0, result.output
    assert keep.read_bytes() == malformed


def test_extract_refuses_catalog_without_raw_and_preserves_project_state(make_xlsx):
    src = make_xlsx(_build, name="orphan-catalog.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / "orphan-catalog.sheetlens"
    catalog_path = proj / "question-ids.json"
    catalog_bytes = catalog_path.read_bytes()
    shutil.rmtree(proj / "structure")

    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 1
    assert "raw.json" in result.output
    assert catalog_path.read_bytes() == catalog_bytes
    assert not (proj / "structure").exists()


def test_reextract_bootstraps_legacy_alias_without_editing_annotations(make_xlsx):
    src = make_xlsx(_build, name="legacy.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / "legacy.sheetlens"
    old_raw = ir.Workbook.model_validate_json(
        (proj / "structure/raw.json").read_text(encoding="utf-8")
    )
    _make_legacy_project(proj)
    annotation, annotation_bytes = _write_legacy_annotation(proj)

    _insert_sheet_first(src, "新規入力")
    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 0, result.output
    catalog = QuestionIdCatalog.model_validate_json(
        (proj / "question-ids.json").read_text(encoding="utf-8")
    )
    target = catalog.legacy_aliases["q-001"]
    assert catalog.questions[target].rule == "sheet_role"
    assert catalog.questions[target].sheet == "見積入力"
    assert catalog.legacy_source_sha256 == old_raw.sha256
    new_raw = ir.Workbook.model_validate_json(
        (proj / "structure/raw.json").read_text(encoding="utf-8")
    )
    assert catalog.source_sha256 == new_raw.sha256
    assert annotation.read_bytes() == annotation_bytes


def test_changed_legacy_questions_prevents_alias_and_preserves_annotations(make_xlsx):
    src = make_xlsx(_build, name="changed-questions.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / "changed-questions.sheetlens"
    _make_legacy_project(proj)
    questions_path = proj / "questions.md"
    questions_path.write_text(
        questions_path.read_text(encoding="utf-8").replace("シート", "改変シート", 1),
        encoding="utf-8",
    )
    annotation, annotation_bytes = _write_legacy_annotation(proj)

    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 0, result.output
    catalog = QuestionIdCatalog.model_validate_json(
        (proj / "question-ids.json").read_text(encoding="utf-8")
    )
    resolution = resolve_answered_ids(["q-001"], catalog)
    assert catalog.legacy_aliases == {}
    assert catalog.legacy_source_sha256 is None
    assert [diagnostic.model_dump() for diagnostic in resolution.diagnostics] == [
        {"kind": "unresolved", "question_id": "q-001", "current_id": None}
    ]
    assert annotation.read_bytes() == annotation_bytes


def test_second_reextract_does_not_repoint_legacy_alias(make_xlsx):
    src = make_xlsx(_build, name="alias-history.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / "alias-history.sheetlens"
    _make_legacy_project(proj)
    annotation, annotation_bytes = _write_legacy_annotation(proj)
    _insert_sheet_first(src, "追加1")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    first_catalog = QuestionIdCatalog.model_validate_json(
        (proj / "question-ids.json").read_text(encoding="utf-8")
    )
    original_target = first_catalog.legacy_aliases["q-001"]
    first_bootstrap_source = first_catalog.legacy_source_sha256
    assert first_bootstrap_source is not None

    _insert_sheet_first(src, "追加2")
    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 0, result.output
    second_catalog = QuestionIdCatalog.model_validate_json(
        (proj / "question-ids.json").read_text(encoding="utf-8")
    )
    assert second_catalog.legacy_aliases["q-001"] == original_target
    assert second_catalog.questions[original_target].sheet == "見積入力"
    assert second_catalog.legacy_source_sha256 == first_bootstrap_source
    assert annotation.read_bytes() == annotation_bytes


def test_reextract_rejects_catalog_source_tampering_before_deleting_structure(make_xlsx):
    src = make_xlsx(_build, name="tampered-catalog.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / "tampered-catalog.sheetlens"
    annotation, annotation_bytes = _write_legacy_annotation(proj)
    raw_bytes = (proj / "structure/raw.json").read_bytes()
    catalog_path = proj / "question-ids.json"
    catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    catalog["source_sha256"] = "f" * 64
    catalog_path.write_text(json.dumps(catalog), encoding="utf-8")

    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 1
    assert "質問 ID エラー:" in result.output
    assert "Traceback" not in result.output
    assert (proj / "structure/raw.json").read_bytes() == raw_bytes
    assert annotation.read_bytes() == annotation_bytes


def test_reextract_identity_error_preserves_entire_structure_and_annotations(make_xlsx):
    src = make_xlsx(_build, name="ambiguous.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / "ambiguous.sheetlens"
    annotation, annotation_bytes = _write_legacy_annotation(proj)
    structure_bytes = _structure_bytes(proj)
    wb = openpyxl.load_workbook(src)
    duplicate = DataValidation(type="list", formula1='"通常,特急,保留"')
    duplicate.add("C5")
    wb["見積入力"].add_data_validation(duplicate)
    wb.save(src)

    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 1
    assert "質問 ID エラー:" in result.output
    assert "Traceback" not in result.output
    assert _structure_bytes(proj) == structure_bytes
    assert annotation.read_bytes() == annotation_bytes


def test_reextract_rejects_tampered_current_ids_before_deleting_structure(make_xlsx):
    src = make_xlsx(_build, name="tampered-current-ids.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / "tampered-current-ids.sheetlens"
    annotation, annotation_bytes = _write_legacy_annotation(proj)
    structure_bytes = _structure_bytes(proj)
    catalog_path = proj / "question-ids.json"
    catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    catalog["current_ids"] = catalog["current_ids"][1:]
    catalog_path.write_text(json.dumps(catalog), encoding="utf-8")

    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 1
    assert "質問 ID エラー:" in result.output
    assert _structure_bytes(proj) == structure_bytes
    assert annotation.read_bytes() == annotation_bytes


def test_reextract_rejects_tampered_current_entry_before_deleting_structure(make_xlsx):
    src = make_xlsx(_build, name="tampered-current-entry.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / "tampered-current-entry.sheetlens"
    annotation, annotation_bytes = _write_legacy_annotation(proj)
    structure_bytes = _structure_bytes(proj)
    catalog_path = proj / "question-ids.json"
    catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    question_id = catalog["current_ids"][0]
    catalog["questions"][question_id]["text"] += " 改ざん"
    catalog_path.write_text(json.dumps(catalog), encoding="utf-8")

    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 1
    assert "質問 ID エラー:" in result.output
    assert _structure_bytes(proj) == structure_bytes
    assert annotation.read_bytes() == annotation_bytes


def test_reextract_removes_stale_structure_files(make_xlsx):
    src = make_xlsx(_build, name="b.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / "b.sheetlens"
    stale = proj / "structure" / "sheet-消えたシート.md"
    stale.write_text("stale", encoding="utf-8")
    keep = proj / "annotations" / "残す.yaml"
    keep.write_text("sheet: 見積入力\n", encoding="utf-8")
    unknown = proj / "user-notes.txt"
    unknown.write_bytes(b"keep unknown\x00")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    assert not stale.exists()
    assert keep.exists()
    assert unknown.read_bytes() == b"keep unknown\x00"


def test_extract_refuses_foreign_structure_dir(make_xlsx, tmp_path):
    src = make_xlsx(_build, name="c.xlsx")
    out = tmp_path / "userproj"
    (out / "structure").mkdir(parents=True)
    user_file = out / "structure" / "main.c"
    user_file.write_text("int main(){}", encoding="utf-8")
    result = runner.invoke(app, ["extract", str(src), "-o", str(out)])
    assert result.exit_code == 1
    assert user_file.exists()
    assert "raw.json" in result.output


def test_extract_rejects_broken_file(tmp_path):
    bad = tmp_path / "broken.xlsx"
    bad.write_bytes(b"not a zip")
    result = runner.invoke(app, ["extract", str(bad)])
    assert result.exit_code == 1
    assert "読めません" in result.output


@pytest.mark.parametrize("failure_point", ["write", "validate"])
def test_reextract_staging_failure_preserves_old_project(
    make_xlsx, monkeypatch, failure_point
):
    src = make_xlsx(_build, name=f"stage-{failure_point}.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / f"stage-{failure_point}.sheetlens"
    annotation = proj / "annotations" / "keep.yaml"
    annotation.write_bytes(b"keep: \xff\x00")
    (proj / "unknown.txt").write_text("keep", encoding="utf-8")
    before = _project_bytes(proj)

    target = (
        "_write_extracted_project"
        if failure_point == "write"
        else "_validate_staged_project"
    )
    original = getattr(pipeline, target)

    def fail(*args, **kwargs):
        original(*args, **kwargs)
        raise OSError("injected staging failure")

    monkeypatch.setattr(pipeline, target, fail)
    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 1
    assert "I/Oエラー" in result.output
    assert _project_bytes(proj) == before
    stage, backup, lock = pipeline._transaction_paths(proj)
    assert not stage.exists() and not backup.exists() and not lock.exists()


@pytest.mark.parametrize("failure_point", ["first", "second"])
def test_reextract_rename_failure_rolls_back(
    make_xlsx, monkeypatch, failure_point
):
    src = make_xlsx(_build, name=f"rename-{failure_point}.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / f"rename-{failure_point}.sheetlens"
    before = _project_bytes(proj)
    stage, backup, lock = pipeline._transaction_paths(proj)
    original_replace = Path.replace

    def fail_replace(self, target):
        if (failure_point == "first" and self == proj) or (
            failure_point == "second" and self == stage
        ):
            raise OSError(f"injected {failure_point} rename failure")
        return original_replace(self, target)

    monkeypatch.setattr(Path, "replace", fail_replace)
    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 1
    assert "I/Oエラー" in result.output
    assert _project_bytes(proj) == before
    assert not stage.exists() and not backup.exists() and not lock.exists()


def test_reextract_rollback_failure_preserves_recovery_paths(make_xlsx, monkeypatch):
    src = make_xlsx(_build, name="rollback-failure.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / "rollback-failure.sheetlens"
    stage, backup, lock = pipeline._transaction_paths(proj)
    original_replace = Path.replace

    def fail_replace(self, target):
        if self in {stage, backup} and target == proj:
            raise OSError("injected swap or rollback failure")
        return original_replace(self, target)

    monkeypatch.setattr(Path, "replace", fail_replace)
    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 1
    assert "復旧エラー" in result.output
    assert str(stage) in result.output and str(backup) in result.output
    assert stage.exists() and backup.exists() and lock.exists()
    assert not proj.exists()


def test_reextract_cleanup_failure_reports_committed_project(make_xlsx, monkeypatch):
    src = make_xlsx(_build, name="cleanup-failure.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / "cleanup-failure.sheetlens"
    _insert_sheet_first(src, "新規")
    stage, backup, lock = pipeline._transaction_paths(proj)
    original_rmtree = shutil.rmtree

    def fail_backup_cleanup(path, *args, **kwargs):
        if Path(path) == backup:
            raise OSError("injected cleanup failure")
        return original_rmtree(path, *args, **kwargs)

    monkeypatch.setattr(pipeline.shutil, "rmtree", fail_backup_cleanup)
    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 0
    assert "警告（後処理）" in result.output
    assert "生成しました" in result.output
    raw = ir.Workbook.model_validate_json(
        (proj / "structure" / "raw.json").read_text(encoding="utf-8")
    )
    assert raw.sheets[0].name == "新規"
    assert backup.exists() and not stage.exists() and not lock.exists()


def test_new_project_final_rename_failure_leaves_no_partial_project(
    make_xlsx, tmp_path, monkeypatch
):
    src = make_xlsx(_build, name="new-rename-failure.xlsx")
    proj = tmp_path / "new-output.sheetlens"
    stage, backup, lock = pipeline._transaction_paths(proj)
    original_replace = Path.replace

    def fail_replace(self, target):
        if self == stage and target == proj:
            raise OSError("injected new project rename failure")
        return original_replace(self, target)

    monkeypatch.setattr(Path, "replace", fail_replace)
    result = runner.invoke(app, ["extract", str(src), "--out", str(proj)])

    assert result.exit_code == 1
    assert "I/Oエラー" in result.output
    assert not proj.exists() and not stage.exists() and not backup.exists() and not lock.exists()


def test_extract_does_not_remove_existing_transaction_lock(make_xlsx):
    src = make_xlsx(_build, name="locked.xlsx")
    proj = src.parent / "locked.sheetlens"
    _, _, lock = pipeline._transaction_paths(proj)
    lock.write_text("pid=other\n", encoding="utf-8")

    result = runner.invoke(app, ["extract", str(src)])

    assert result.exit_code == 1
    assert str(lock) in result.output and "手動で復旧" in result.output
    assert lock.read_text(encoding="utf-8") == "pid=other\n"
    assert not proj.exists()


def test_extract_rejects_stale_transaction_and_managed_symlink(make_xlsx):
    src = make_xlsx(_build, name="transaction-safety.xlsx")
    assert runner.invoke(app, ["extract", str(src)]).exit_code == 0
    proj = src.parent / "transaction-safety.sheetlens"
    before = _project_bytes(proj)
    stage, backup, lock = pipeline._transaction_paths(proj)
    stage.mkdir()

    stale = runner.invoke(app, ["extract", str(src)])

    assert stale.exit_code == 1
    assert str(stage) in stale.output and "手動復旧" in stale.output
    assert _project_bytes(proj) == before
    assert stage.exists() and not backup.exists() and not lock.exists()

    stage.rmdir()
    manifest = proj / "manifest.json"
    target = proj / "manifest-target.json"
    manifest.replace(target)
    manifest.symlink_to(target.name)

    unsafe = runner.invoke(app, ["extract", str(src)])

    assert unsafe.exit_code == 1
    assert str(manifest) in unsafe.output and "symlink" in unsafe.output
