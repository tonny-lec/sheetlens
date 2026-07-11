import openpyxl
import pytest
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
    Rule,
)
from openpyxl.styles import Border, Color, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation

from sheetlens.reader.features import read_conditional_formats, read_validations
from sheetlens.reader.workbook import read_workbook


def _add_list_validation(ws, target: str, formula: str) -> None:
    dv = DataValidation(type="list", formula1=formula)
    dv.add(target)
    ws.add_data_validation(dv)


def _rules_by_range(workbook, sheet_name: str = "入力"):
    sheet = next(sheet for sheet in workbook.sheets if sheet.name == sheet_name)
    return {rule.ranges[0]: rule for rule in sheet.validations}


def _build(wb):
    ws = wb.active
    ws.title = "入力"
    master = wb.create_sheet("区分マスタ")
    for i, v in enumerate(["通常", "特急"], start=2):
        master[f"A{i}"] = v
    dv_inline = DataValidation(type="list", formula1='"はい,いいえ"')
    dv_inline.add("B2")
    ws.add_data_validation(dv_inline)
    dv_ref = DataValidation(type="list", formula1="=区分マスタ!$A$2:$A$3")
    dv_ref.add("C5")
    ws.add_data_validation(dv_ref)
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add("F1:F9", CellIsRule(operator="lessThan", formula=["0"], fill=red))


def test_validations_and_conditional_formats(make_xlsx):
    sheet = read_workbook(make_xlsx(_build)).sheets[0]
    rules = {r.ranges[0]: r for r in sheet.validations}
    assert rules["B2"].choices == ["はい", "いいえ"]
    assert rules["C5"].choices == ["通常", "特急"]
    cf = sheet.conditional_formats[0]
    assert cf.range == "F1:F9"
    assert cf.rule_type == "cellIs"
    assert cf.operator == "lessThan"
    assert cf.formula == "0"


def test_conditional_formats_preserve_all_formulas_ranges_and_dxf(make_xlsx):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        thin = Side(style="thin", color="FF00FF00")
        ws.conditional_formatting.add(
            "A1:A3 C1:C3",
            CellIsRule(
                operator="between",
                formula=["1", "10"],
                stopIfTrue=True,
                font=Font(bold=True, color="FF0000FF"),
                fill=red,
                border=Border(left=thin),
            ),
        )
        ws.conditional_formatting.add("D1:D3", FormulaRule(formula=["D1>0", "D1<10"]))

    workbook = read_workbook(make_xlsx(build))
    rules = workbook.sheets[0].conditional_formats

    assert rules[0].range == "A1:A3 C1:C3"
    assert rules[0].formulas == ["1", "10"]
    assert rules[0].operator == "between"
    assert rules[0].stop_if_true is True
    assert rules[0].dxf is not None
    assert rules[0].dxf.tag.endswith("dxf")
    child_tags = [child.tag.rsplit("}", 1)[-1] for child in rules[0].dxf.children]
    assert child_tags == ["font", "fill", "border"]
    assert any(
        descendant.attributes.get("rgb") == "FF0000FF"
        for child in rules[0].dxf.children
        for descendant in child.children
    )
    assert rules[1].formulas == ["D1>0", "D1<10"]
    assert workbook.extraction_gaps == []


def test_conditional_format_dxf_failure_adds_gap_and_keeps_rule(monkeypatch):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入力"
    ws.conditional_formatting.add(
        "A1",
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            fill=PatternFill(fill_type="solid", fgColor="FFFF0000"),
        ),
    )
    rule = next(iter(ws.conditional_formatting)).rules[0]
    monkeypatch.setattr(
        type(rule.dxf),
        "to_tree",
        lambda self: (_ for _ in ()).throw(RuntimeError("broken dxf")),
    )
    gaps = ["既存gap"]

    extracted = read_conditional_formats(ws, extraction_gaps=gaps)

    assert extracted[0].formulas == ["0"]
    assert extracted[0].dxf is None
    assert gaps == [
        "既存gap",
        "入力: 条件付き書式 A1 を完全に抽出できません (type=cellIs; reason=invalid_dxf)",
    ]


def test_conditional_format_rule_failure_keeps_later_rule():
    class BrokenRule:
        type = "expression"

        @property
        def formula(self):
            raise RuntimeError("broken formula")

    class GoodRule:
        type = "expression"
        formula = ["A1>0"]
        operator = None
        stopIfTrue = False
        dxf = None

    class Format:
        sqref = "A1:A2"
        rules = [BrokenRule(), GoodRule()]

    class Worksheet:
        title = "入力"
        conditional_formatting = [Format()]

    gaps: list[str] = []
    extracted = read_conditional_formats(Worksheet(), extraction_gaps=gaps)

    assert [(rule.rule_type, rule.formulas) for rule in extracted] == [
        ("expression", []),
        ("expression", ["A1>0"]),
    ]
    assert gaps == [
        "入力: 条件付き書式 A1:A2 を完全に抽出できません (type=expression; reason=extraction_error)"
    ]


def test_conditional_format_visual_payloads_survive_workbook_round_trip(make_xlsx):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        ws.conditional_formatting.add(
            "A1:A3",
            ColorScaleRule(
                start_type="min",
                start_color=Color(theme=4, tint=0.25),
                mid_type="percentile",
                mid_value=50,
                mid_color=Color(indexed=7),
                end_type="max",
                end_color="FFFF0000",
            ),
        )
        data_bar = DataBarRule(
            start_type="min",
            end_type="max",
            color="FF638EC6",
            showValue=False,
            minLength=5,
            maxLength=95,
        )
        data_bar.dataBar.color = Color(auto=True)
        ws.conditional_formatting.add("B1:B3", data_bar)
        ws.conditional_formatting.add(
            "C1:C3",
            IconSetRule(
                icon_style="3TrafficLights1",
                type="percent",
                values=[0, 33, 67],
                showValue=False,
                percent=True,
                reverse=True,
            ),
        )

    workbook = read_workbook(make_xlsx(build))
    rules = workbook.sheets[0].conditional_formats

    scale = rules[0].color_scale
    assert scale is not None
    assert [(v.type, v.value, v.gte) for v in scale.conditions] == [
        ("min", None, None),
        ("percentile", 50.0, None),
        ("max", None, None),
    ]
    assert [(c.type, c.value, c.tint) for c in scale.colors] == [
        ("theme", 4, 0.25),
        ("indexed", 7, 0.0),
        ("rgb", "FFFF0000", 0.0),
    ]
    bar = rules[1].data_bar
    assert bar is not None
    assert [value.type for value in bar.conditions] == ["min", "max"]
    assert (bar.color.type, bar.color.value) == ("auto", True)
    assert (bar.show_value, bar.min_length, bar.max_length) == (False, 5, 95)
    icons = rules[2].icon_set
    assert icons is not None
    assert icons.icon_style == "3TrafficLights1"
    assert [value.value for value in icons.conditions] == [0.0, 33.0, 67.0]
    assert (icons.show_value, icons.percent, icons.reverse) == (False, True, True)
    assert workbook.extraction_gaps == []


@pytest.mark.parametrize(
    ("rule_type", "payload_name", "payload", "reason"),
    [
        ("colorScale", "colorScale", None, "missing_color_scale"),
        ("colorScale", "colorScale", object(), "invalid_color_scale"),
        ("dataBar", "dataBar", None, "missing_data_bar"),
        ("dataBar", "dataBar", object(), "invalid_data_bar"),
        ("iconSet", "iconSet", None, "missing_icon_set"),
        ("iconSet", "iconSet", object(), "invalid_icon_set"),
    ],
)
def test_conditional_format_malformed_payload_reason(
    rule_type, payload_name, payload, reason
):
    class Rule:
        type = rule_type
        formula = []
        operator = None
        stopIfTrue = False
        dxf = None

    setattr(Rule, payload_name, payload)

    class Format:
        sqref = "A1"
        rules = [Rule()]

    class Worksheet:
        title = "入力"
        conditional_formatting = [Format()]

    gaps: list[str] = []
    extracted = read_conditional_formats(Worksheet(), extraction_gaps=gaps)

    assert len(extracted) == 1
    assert gaps == [
        f"入力: 条件付き書式 A1 を完全に抽出できません (type={rule_type}; reason={reason})"
    ]


def test_conditional_format_unsupported_rule_keeps_common_fields_and_one_gap():
    class Rule:
        type = "top10"
        formula = ["A1"]
        operator = None
        stopIfTrue = True
        dxf = None

    class Format:
        sqref = "A1:A3"
        rules = [Rule()]

    class Worksheet:
        title = "入力"
        conditional_formatting = [Format()]

    gaps: list[str] = []
    extracted = read_conditional_formats(Worksheet(), extraction_gaps=gaps)

    assert extracted[0].formulas == ["A1"]
    assert extracted[0].stop_if_true is True
    assert len(gaps) == 1
    assert "reason=unsupported_type" in gaps[0]


def test_conditional_format_dxf_failure_keeps_visual_payload(monkeypatch):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入力"
    rule = ColorScaleRule(
        start_type="min",
        start_color="FF00FF00",
        end_type="max",
        end_color="FFFF0000",
    )
    rule.dxf = DifferentialStyle(font=Font(bold=True))
    ws.conditional_formatting.add("A1:A3", rule)
    monkeypatch.setattr(
        type(rule.dxf),
        "to_tree",
        lambda self: (_ for _ in ()).throw(RuntimeError("broken dxf")),
    )
    gaps: list[str] = []

    extracted = read_conditional_formats(ws, extraction_gaps=gaps)

    assert extracted[0].dxf is None
    assert extracted[0].color_scale is not None
    assert len(gaps) == 1
    assert "reason=invalid_dxf" in gaps[0]


def test_conditional_format_rule_gap_reaches_workbook(make_xlsx):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        ws.conditional_formatting.add("A1", Rule(type="colorScale"))

    workbook = read_workbook(make_xlsx(build))

    assert len(workbook.sheets[0].conditional_formats) == 1
    assert workbook.extraction_gaps == [
        "入力: 条件付き書式 A1 を完全に抽出できません "
        "(type=colorScale; reason=missing_color_scale)"
    ]


def test_resolves_workbook_name_case_insensitively_and_quoted_sheet(make_xlsx):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        master = wb.create_sheet("O'Brien")
        master["A2"] = "通常"
        master["A3"] = "特急"
        wb.defined_names.add(
            DefinedName("Choices", attr_text="'O''Brien'!$A$2:$A$3")
        )
        _add_list_validation(ws, "B2", "=choices")
        _add_list_validation(ws, "C2", "='O''Brien'!$A$2:$A$3")

    workbook = read_workbook(make_xlsx(build))
    rules = _rules_by_range(workbook)
    assert rules["B2"].choices == ["通常", "特急"]
    assert rules["C2"].choices == ["通常", "特急"]
    assert workbook.extraction_gaps == []


def test_resolves_current_sheet_range_and_distinguishes_valid_empty_range(make_xlsx):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        ws["D2"] = "赤"
        ws["D3"] = "青"
        wb.create_sheet("空マスタ")
        wb.defined_names.add(
            DefinedName("EmptyChoices", attr_text="'空マスタ'!$A$1:$A$2")
        )
        _add_list_validation(ws, "B2", "=$D$2:$D$3")
        _add_list_validation(ws, "C2", "=EmptyChoices")

    workbook = read_workbook(make_xlsx(build))
    rules = _rules_by_range(workbook)
    assert rules["B2"].choices == ["赤", "青"]
    assert rules["C2"].choices == []
    assert workbook.extraction_gaps == []


def test_sheet_local_name_shadows_workbook_name_and_other_sheet_falls_back(make_xlsx):
    def build(wb):
        input_ws = wb.active
        input_ws.title = "入力"
        other_ws = wb.create_sheet("別シート")
        master = wb.create_sheet("共通マスタ")
        master["A2"] = "共通1"
        master["A3"] = "共通2"
        input_ws["D2"] = "ローカル1"
        input_ws["D3"] = "ローカル2"
        wb.defined_names.add(
            DefinedName("Choices", attr_text="'共通マスタ'!$A$2:$A$3")
        )
        input_ws.defined_names.add(
            DefinedName("cHoIcEs", attr_text="$D$2:$D$3")
        )
        _add_list_validation(input_ws, "B2", "=CHOICES")
        _add_list_validation(other_ws, "B2", "=choices")

    workbook = read_workbook(make_xlsx(build))
    input_rules = _rules_by_range(workbook, "入力")
    other_rules = _rules_by_range(workbook, "別シート")
    assert input_rules["B2"].choices == ["ローカル1", "ローカル2"]
    assert other_rules["B2"].choices == ["共通1", "共通2"]
    assert workbook.extraction_gaps == []


@pytest.mark.parametrize(
    ("formula", "reason"),
    [
        ("=MissingChoices", "name_not_found"),
        ('=INDIRECT("D2:D3")', "unsupported_indirect"),
        ("=OFFSET(D2,0,0,2,1)", "unsupported_offset"),
        ("=SUM(A1:A2)", "unsupported_reference"),
        (
            "='入力'!$D$2:$D$3,'入力'!$E$2:$E$3",
            "unsupported_reference",
        ),
    ],
)
def test_unresolved_list_source_keeps_rule_and_adds_one_gap(
    make_xlsx, formula, reason
):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        _add_list_validation(ws, "C2", formula)
        dv = DataValidation(type="list", formula1=formula)
        dv.add("D2")
        dv.add("B2")
        ws.add_data_validation(dv)

    workbook = read_workbook(make_xlsx(build))
    sheet = workbook.sheets[0]
    assert len(sheet.validations) == 2
    assert all(rule.formula1 == formula for rule in sheet.validations)
    assert all(rule.choices == [] for rule in sheet.validations)
    assert workbook.extraction_gaps == [
        f"入力: 入力規則 C2 の選択肢を解決できません "
        f"(formula1={formula!r}; reason={reason})",
        f"入力: 入力規則 B2, D2 の選択肢を解決できません "
        f"(formula1={formula!r}; reason={reason})",
    ]


def test_quoted_sheet_name_with_comma_resolves_direct_and_defined_ranges(make_xlsx):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        master = wb.create_sheet("Sales, 2026")
        master["A1"] = "one"
        master["A2"] = "two"
        reference = "'Sales, 2026'!$A$1:$A$2"
        wb.defined_names.add(DefinedName("CommaChoices", attr_text=reference))
        _add_list_validation(ws, "B2", f"={reference}")
        _add_list_validation(ws, "C2", "=CommaChoices")

    workbook = read_workbook(make_xlsx(build))
    rules = _rules_by_range(workbook)
    assert rules["B2"].choices == ["one", "two"]
    assert rules["C2"].choices == ["one", "two"]
    assert workbook.extraction_gaps == []


@pytest.mark.parametrize(
    "reference",
    [
        "Sheet1:Sheet2!$A$1",
        "[Book.xlsx]Sheet1!$A$1",
    ],
)
def test_unsupported_3d_and_external_references_add_gaps_after_round_trip(
    make_xlsx, reference
):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        wb.create_sheet("Sheet1")
        wb.create_sheet("Sheet2")
        wb.defined_names.add(
            DefinedName("UnsupportedChoices", attr_text=reference)
        )
        _add_list_validation(ws, "B2", f"={reference}")
        _add_list_validation(ws, "C2", "=UnsupportedChoices")

    workbook = read_workbook(make_xlsx(build))
    rules = _rules_by_range(workbook)
    assert rules["B2"].choices == []
    assert rules["C2"].choices == []
    assert workbook.extraction_gaps == [
        "入力: 入力規則 B2 の選択肢を解決できません "
        f"(formula1={'=' + reference!r}; reason=unsupported_reference)",
        "入力: 入力規則 C2 の選択肢を解決できません "
        "(formula1='=UnsupportedChoices'; reason=unsupported_reference)",
    ]


def test_trailing_text_in_qualified_range_adds_invalid_range_gaps(make_xlsx):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        ws["A1"] = "one"
        ws["A2"] = "two"
        malformed = "'入力'!$A$1:$A$2garbage"
        wb.defined_names.add(DefinedName("TrailingGarbage", attr_text=malformed))
        _add_list_validation(ws, "B2", f"={malformed}")
        _add_list_validation(ws, "C2", "=TrailingGarbage")

    workbook = read_workbook(make_xlsx(build))
    rules = _rules_by_range(workbook)
    assert rules["B2"].choices == []
    assert rules["C2"].choices == []
    assert workbook.extraction_gaps == [
        "入力: 入力規則 B2 の選択肢を解決できません "
        '(formula1="=\'入力\'!$A$1:$A$2garbage"; reason=invalid_range)',
        "入力: 入力規則 C2 の選択肢を解決できません "
        "(formula1='=TrailingGarbage'; reason=invalid_range)",
    ]


def test_unresolved_validation_does_not_discard_valid_rule_on_same_sheet(make_xlsx):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        master = wb.create_sheet("マスタ")
        master["A1"] = "one"
        master["A2"] = "two"
        _add_list_validation(ws, "B2", "='マスタ'!$A$1:$A$2")
        _add_list_validation(ws, "C2", "=MissingChoices")

    workbook = read_workbook(make_xlsx(build))
    sheet = workbook.sheets[0]
    rules = _rules_by_range(workbook)
    assert len(sheet.validations) == 2
    assert rules["B2"].choices == ["one", "two"]
    assert rules["C2"].choices == []
    assert workbook.extraction_gaps == [
        "入力: 入力規則 C2 の選択肢を解決できません "
        "(formula1='=MissingChoices'; reason=name_not_found)"
    ]


@pytest.mark.parametrize(
    ("name", "attr_text", "reason"),
    [
        ("DynamicIndirect", 'INDIRECT("入力!$D$2:$D$3")', "unsupported_indirect"),
        ("DynamicOffset", "OFFSET('入力'!$D$2,0,0,2,1)", "unsupported_offset"),
        ("UnqualifiedGlobal", "$D$2:$D$3", "unsupported_reference"),
        ("MissingSheet", "'存在しない'!$A$1:$A$2", "sheet_not_found"),
        ("BrokenRange", "'入力'!#REF!", "invalid_range"),
        (
            "MultipleAreas",
            "'入力'!$D$2:$D$3,'入力'!$E$2:$E$3",
            "unsupported_reference",
        ),
    ],
)
def test_unsupported_workbook_name_definition_adds_gap(
    make_xlsx, name, attr_text, reason
):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        wb.defined_names.add(DefinedName(name, attr_text=attr_text))
        _add_list_validation(ws, "B2", f"={name}")

    workbook = read_workbook(make_xlsx(build))
    rule = workbook.sheets[0].validations[0]
    assert rule.choices == []
    assert rule.formula1 == f"={name}"
    formula = f"={name}"
    assert workbook.extraction_gaps == [
        "入力: 入力規則 B2 の選択肢を解決できません "
        f"(formula1={formula!r}; reason={reason})"
    ]


def test_invalid_local_name_shadows_valid_workbook_name(make_xlsx):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        master = wb.create_sheet("共通マスタ")
        master["A2"] = "共通1"
        master["A3"] = "共通2"
        wb.defined_names.add(
            DefinedName("Choices", attr_text="'共通マスタ'!$A$2:$A$3")
        )
        ws.defined_names.add(
            DefinedName("choices", attr_text="OFFSET($D$2,0,0,2,1)")
        )
        _add_list_validation(ws, "B2", "=CHOICES")

    workbook = read_workbook(make_xlsx(build))
    assert workbook.sheets[0].validations[0].choices == []
    assert workbook.extraction_gaps[0].endswith("reason=unsupported_offset)")


def test_case_insensitive_duplicate_names_are_ambiguous(make_xlsx):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        first = wb.create_sheet("第一")
        second = wb.create_sheet("第二")
        first["A1"] = "一"
        second["A1"] = "二"
        wb.defined_names.add(
            DefinedName("Choices", attr_text="'第一'!$A$1")
        )
        wb.defined_names.add(
            DefinedName("choices", attr_text="'第二'!$A$1")
        )
        _add_list_validation(ws, "B2", "=CHOICES")

    workbook = read_workbook(make_xlsx(build))
    assert workbook.sheets[0].validations[0].choices == []
    assert workbook.extraction_gaps == [
        "入力: 入力規則 B2 の選択肢を解決できません "
        "(formula1='=CHOICES'; reason=ambiguous_name)"
    ]


def test_read_validations_preserves_old_return_type_and_appends_to_gap_sink(make_xlsx):
    def build(wb):
        ws = wb.active
        ws.title = "入力"
        _add_list_validation(ws, "B2", "=MissingChoices")

    path = make_xlsx(build)
    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True)
    empty_sink: list[str] = []
    existing_sink = ["既存gap"]

    rules = read_validations(
        wb_f["入力"],
        wb_v,
        extraction_gaps=empty_sink,
    )
    read_validations(wb_f["入力"], wb_v, extraction_gaps=existing_sink)
    legacy_rules = read_validations(wb_f["入力"], wb_v)

    assert isinstance(rules, list)
    assert isinstance(legacy_rules, list)
    expected = (
        "入力: 入力規則 B2 の選択肢を解決できません "
        "(formula1='=MissingChoices'; reason=name_not_found)"
    )
    assert empty_sink == [expected]
    assert existing_sink == ["既存gap", expected]
