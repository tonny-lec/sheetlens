from datetime import date, datetime, time, timedelta
from pathlib import Path
import zipfile

import openpyxl
import pytest
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from sheetlens.reader.artifacts import extract_sheet_artifacts
from sheetlens.reader.workbook import (
    _formula_text,
    _has_currency_format,
    _has_leading_zero_format,
    read_workbook,
)


def _build(wb):
    ws = wb.active
    ws.title = "見積入力"
    ws["A1"] = "見積書"
    ws.merge_cells("A1:C1")
    ws["A3"] = "数量"
    ws["B3"] = 5
    ws["C3"] = "=B3*100"
    ws.column_dimensions["D"].hidden = True
    hidden = wb.create_sheet("計算用")
    hidden["A1"] = 1
    hidden.sheet_state = "hidden"


def _artifact_package(base_path: Path, output_path: Path, *, dangling_chart: bool = False) -> Path:
    with zipfile.ZipFile(base_path) as source:
        files = {name: source.read(name) for name in source.namelist()}

    sheet_name = "xl/worksheets/sheet1.xml"
    custom_sheet_name = "xl/worksheets/custom/data.xml"
    sheet_xml = files.pop(sheet_name).decode()
    if "xmlns:r=" not in sheet_xml:
        sheet_xml = sheet_xml.replace(
            "<worksheet ",
            '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
            1,
        )
    sheet_xml = sheet_xml.replace(
        "</worksheet>",
        '<drawing r:id="rIdDrawing"/>'
        '<pivotTableParts count="2">'
        '<pivotTablePart r:id="rIdPivot"/>'
        '<pivotTablePart r:id="rIdPivotCache"/>'
        "</pivotTableParts></worksheet>",
    )
    files[custom_sheet_name] = sheet_xml.encode()
    workbook_rels = files["xl/_rels/workbook.xml.rels"].decode()
    files["xl/_rels/workbook.xml.rels"] = workbook_rels.replace(
        "worksheets/sheet1.xml", "worksheets/custom/data.xml"
    ).encode()

    files["xl/worksheets/custom/_rels/data.xml.rels"] = b"""\
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rIdDrawing" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../../drawings/drawing9.xml"/>
  <Relationship Id="rIdPivot" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" Target="../../pivotTables/pivotTable7.xml"/>
  <Relationship Id="rIdPivotCache" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" Target="../../pivotCache/pivotCacheDefinition1.xml"/>
</Relationships>"""
    chart_rel = "" if dangling_chart else """\
  <Relationship Id="rIdChart" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart" Target="../charts/chart3.xml"/>"""
    files["xl/drawings/_rels/drawing9.xml.rels"] = f"""\
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
{chart_rel}
  <Relationship Id="rIdImage" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/image1.png"/>
</Relationships>""".encode()
    chart_id = "rIdMissing" if dangling_chart else "rIdChart"
    files["xl/drawings/drawing9.xml"] = f"""\
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
 xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
 xmlns:c="http://schemas.openxmlformats.org/drawingml/2006/chart"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <xdr:twoCellAnchor><xdr:graphicFrame><a:graphic><a:graphicData><c:chart r:id="{chart_id}"/></a:graphicData></a:graphic></xdr:graphicFrame></xdr:twoCellAnchor>
  <xdr:oneCellAnchor><xdr:pic><xdr:blipFill><a:blip r:embed="rIdImage"/></xdr:blipFill></xdr:pic></xdr:oneCellAnchor>
  <xdr:absoluteAnchor><xdr:pic><xdr:blipFill><a:blip r:embed="rIdImage"/></xdr:blipFill></xdr:pic></xdr:absoluteAnchor>
  <xdr:twoCellAnchor><xdr:sp/></xdr:twoCellAnchor>
  <xdr:twoCellAnchor><xdr:cxnSp/></xdr:twoCellAnchor>
  <xdr:twoCellAnchor><xdr:grpSp><xdr:sp/></xdr:grpSp></xdr:twoCellAnchor>
</xdr:wsDr>""".encode()
    files["xl/charts/chart3.xml"] = b"<chart/>"
    files["xl/media/image1.png"] = b"image bytes are intentionally not decoded"
    files["xl/pivotTables/pivotTable7.xml"] = b"<pivotTableDefinition/>"
    files["xl/pivotCache/pivotCacheDefinition1.xml"] = b"<pivotCacheDefinition/>"

    with zipfile.ZipFile(output_path, "w") as target:
        for name, data in files.items():
            target.writestr(name, data)
    return output_path


def _read_raw_artifact_package(monkeypatch, base_path: Path, package_path: Path):
    real_load_workbook = openpyxl.load_workbook

    def load_base_workbook(_path, *args, **kwargs):
        return real_load_workbook(base_path, *args, **kwargs)

    monkeypatch.setattr("sheetlens.reader.workbook.openpyxl.load_workbook", load_base_workbook)
    return read_workbook(package_path)


def _rewrite_package(
    source_path: Path,
    output_path: Path,
    replacements: list[tuple[bytes, bytes]],
) -> Path:
    with zipfile.ZipFile(source_path) as source, zipfile.ZipFile(output_path, "w") as target:
        for name in source.namelist():
            data = source.read(name)
            for old, new in replacements:
                data = data.replace(old, new)
            target.writestr(name, data)
    return output_path


def test_read_cells_formulas_merges(make_xlsx):
    wb = read_workbook(make_xlsx(_build))
    assert wb.source_file == "test.xlsx"
    assert len(wb.sha256) == 64
    sheet = wb.sheets[0]
    assert sheet.name == "見積入力"
    assert "A1:C1" in sheet.merged
    cells = {c.ref: c for c in sheet.cells}
    assert cells["A1"].value == "見積書"
    assert cells["B3"].value == 5
    assert cells["C3"].formula == "=B3*100"
    assert "D" in sheet.hidden_cols
    assert wb.sheets[1].hidden is True


def test_read_workbook_records_chart_reused_image_shapes_and_pivot(
    make_xlsx, tmp_path, monkeypatch
):
    base_path = make_xlsx(lambda wb: setattr(wb.active, "title", "見積入力"))
    package_path = _artifact_package(base_path, tmp_path / "artifacts.xlsx")

    wb = _read_raw_artifact_package(monkeypatch, base_path, package_path)

    assert [artifact.model_dump() for artifact in wb.sheets[0].artifacts] == [
        {"type": "chart", "count": 1, "ooxml_parts": ["xl/charts/chart3.xml"]},
        {"type": "image", "count": 2, "ooxml_parts": ["xl/media/image1.png"]},
        {"type": "shape", "count": 3, "ooxml_parts": ["xl/drawings/drawing9.xml"]},
        {"type": "pivot", "count": 1, "ooxml_parts": ["xl/pivotTables/pivotTable7.xml"]},
    ]
    assert not any("pivotCacheDefinition1.xml" in gap for gap in wb.extraction_gaps)
    for artifact_type in ("chart", "image", "shape", "pivot"):
        assert any(
            artifact_type in gap and "詳細" in gap and "未対応" in gap
            for gap in wb.extraction_gaps
        )


def test_dangling_drawing_relationship_keeps_other_artifacts_and_adds_gap(
    make_xlsx, tmp_path
):
    base_path = make_xlsx(lambda wb: setattr(wb.active, "title", "見積入力"))
    package_path = _artifact_package(
        base_path, tmp_path / "dangling.xlsx", dangling_chart=True
    )

    artifacts_by_sheet, gaps = extract_sheet_artifacts(package_path)
    artifacts = {
        artifact.type: artifact for artifact in artifacts_by_sheet["見積入力"]
    }

    assert artifacts["chart"].count == 1
    assert artifacts["chart"].ooxml_parts == []
    assert artifacts["image"].count == 2
    assert artifacts["shape"].count == 3
    assert artifacts["pivot"].count == 1
    assert any(
        "見積入力" in gap and "rIdMissing" in gap and "未解決" in gap
        for gap in gaps
    )


def test_arbitrary_namespaces_and_suffix_matching_relationship_types_are_not_artifacts(
    make_xlsx, tmp_path
):
    base_path = make_xlsx(lambda wb: setattr(wb.active, "title", "見積入力"))
    package_path = _artifact_package(base_path, tmp_path / "artifacts.xlsx")
    hostile_path = _rewrite_package(
        package_path,
        tmp_path / "hostile.xlsx",
        [
            (
                b'<drawing r:id="rIdDrawing"/><pivotTableParts count="2"><pivotTablePart r:id="rIdPivot"/><pivotTablePart r:id="rIdPivotCache"/></pivotTableParts>',
                b'<evil:drawing xmlns:evil="urn:evil/elements" r:id="rIdDrawing"/>',
            ),
            (
                b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing",
                b"urn:evil/drawing",
            ),
            (
                b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/chart",
                b"urn:evil/chart",
            ),
            (
                b"http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                b"urn:evil/elements",
            ),
            (
                b"http://schemas.openxmlformats.org/drawingml/2006/chart",
                b"urn:evil/elements",
            ),
        ],
    )

    artifacts_by_sheet, gaps = extract_sheet_artifacts(hostile_path)

    assert artifacts_by_sheet["見積入力"] == []
    assert any("未対応" in gap or "unknown" in gap for gap in gaps)


def test_official_strict_namespaces_and_relationship_types_are_supported(
    make_xlsx, tmp_path
):
    base_path = make_xlsx(lambda wb: setattr(wb.active, "title", "見積入力"))
    package_path = _artifact_package(base_path, tmp_path / "transitional.xlsx")
    strict_path = _rewrite_package(
        package_path,
        tmp_path / "strict.xlsx",
        [
            (
                b"http://schemas.openxmlformats.org/officeDocument/2006/relationships",
                b"http://purl.oclc.org/ooxml/officeDocument/relationships",
            ),
            (
                b"http://schemas.openxmlformats.org/spreadsheetml/2006/main",
                b"http://purl.oclc.org/ooxml/spreadsheetml/main",
            ),
            (
                b"http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                b"http://purl.oclc.org/ooxml/drawingml/spreadsheetDrawing",
            ),
            (
                b"http://schemas.openxmlformats.org/drawingml/2006/main",
                b"http://purl.oclc.org/ooxml/drawingml/main",
            ),
            (
                b"http://schemas.openxmlformats.org/drawingml/2006/chart",
                b"http://purl.oclc.org/ooxml/drawingml/chart",
            ),
        ],
    )

    artifacts_by_sheet, _ = extract_sheet_artifacts(strict_path)

    assert [artifact.type for artifact in artifacts_by_sheet["見積入力"]] == [
        "chart",
        "image",
        "shape",
        "pivot",
    ]


def test_read_workbook_integrates_real_openpyxl_chart(make_xlsx):
    def build(wb):
        ws = wb.active
        ws.title = "グラフ"
        ws.append(["月", "売上"])
        ws.append([1, 10])
        ws.append([2, 20])
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
        ws.add_chart(chart, "D2")

    wb = read_workbook(make_xlsx(build, name="chart.xlsx"))

    assert [artifact.model_dump() for artifact in wb.sheets[0].artifacts] == [
        {"type": "chart", "count": 1, "ooxml_parts": ["xl/charts/chart1.xml"]}
    ]
    assert any("chart" in gap and "詳細" in gap and "未対応" in gap for gap in wb.extraction_gaps)


def test_read_workbook_propagates_unexpected_artifact_scanner_error(
    make_xlsx, monkeypatch
):
    path = make_xlsx(lambda wb: None)

    def fail(_path):
        raise RuntimeError("scanner invariant broken")

    monkeypatch.setattr("sheetlens.reader.workbook.extract_sheet_artifacts", fail)

    with pytest.raises(RuntimeError, match="scanner invariant broken"):
        read_workbook(path)


def test_formula_text_unwraps_known_types():
    assert _formula_text("=A1*2") == "=A1*2"
    assert _formula_text(ArrayFormula("A1:A3", "=SUM(B1:B3)")) == "=SUM(B1:B3)"
    assert _formula_text(DataTableFormula("A1")) is None


def test_bounded_currency_and_leading_zero_format_classification():
    assert _has_currency_format("[$USD] #,##0.00") is True
    assert _has_currency_format("[$USD-409] #,##0.00") is True
    assert _has_currency_format("[$-409]0.00") is False
    assert _has_leading_zero_format("00000") is True
    assert _has_leading_zero_format("000-000") is True
    assert _has_leading_zero_format("0/00") is False
    assert _has_leading_zero_format("00E+00") is False
    assert _has_leading_zero_format("0 0") is False


def _build_display_semantics(wb):
    ws = wb.active
    ws["A1"] = 0.125
    ws["A1"].number_format = "0.00%"
    ws["A2"] = 1234.5
    ws["A2"].number_format = "¥#,##0.00"
    ws["A3"] = date(2026, 7, 11)
    ws["A3"].number_format = "yyyy-mm-dd"
    ws["A4"] = time(14, 30, 5)
    ws["A4"].number_format = "hh:mm:ss"
    ws["A5"] = datetime(2026, 7, 11, 14, 30, 5)
    ws["A5"].number_format = "yyyy-mm-dd hh:mm:ss"
    ws["A6"] = timedelta(hours=27, minutes=5)
    ws["A6"].number_format = "[h]:mm:ss"
    ws["A7"] = 123
    ws["A7"].number_format = "00000"
    ws["A8"] = "00123"
    ws["A9"] = "#DIV/0!"
    ws["A10"] = "=1/0"
    ws["A11"] = 12.5
    ws["A11"].number_format = '"%"0.00'
    ws["A12"] = 12.5
    ws["A12"].number_format = r"\%0.00"
    ws["A13"] = "=1/4"
    ws["A13"].number_format = "0.00%"
    ws["A14"] = "=1000"
    ws["A14"].number_format = "$#,##0"
    ws["A15"] = "=DATE(2026,7,11)"
    ws["A15"].number_format = "yyyy-mm-dd"
    ws["A16"] = "=123"
    ws["A16"].number_format = "00000"
    ws["A17"] = date(2026, 7, 11)
    ws["A17"].number_format = "YYYY-MM-DD"


def test_read_cell_types_number_formats_and_display_semantics(make_xlsx):
    wb = read_workbook(make_xlsx(_build_display_semantics))
    restored = type(wb).model_validate_json(wb.model_dump_json())
    cells = {cell.ref: cell for cell in restored.sheets[0].cells}

    assert (cells["A1"].value_type, cells["A1"].number_format, cells["A1"].display_semantics) == (
        "number",
        "0.00%",
        "percentage",
    )
    assert (cells["A2"].value_type, cells["A2"].display_semantics) == ("number", "currency")
    assert (cells["A3"].value_type, cells["A3"].display_semantics) == ("date", "date")
    assert (cells["A4"].value_type, cells["A4"].display_semantics) == ("time", "time")
    assert (cells["A5"].value_type, cells["A5"].display_semantics) == (
        "datetime",
        "datetime",
    )
    assert (cells["A6"].value_type, cells["A6"].display_semantics) == (
        "duration",
        "duration",
    )
    assert (cells["A7"].value_type, cells["A7"].display_semantics) == (
        "number",
        "leading_zero",
    )
    assert (cells["A8"].value_type, cells["A8"].display_semantics) == (
        "string",
        "leading_zero",
    )
    assert (cells["A9"].value_type, cells["A9"].display_semantics) == ("error", "error")
    assert cells["A10"].formula == "=1/0"
    assert cells["A10"].value is None
    assert cells["A10"].value_type is None
    assert cells["A10"].number_format == "General"
    assert cells["A10"].display_semantics is None
    assert cells["A11"].display_semantics is None
    assert cells["A12"].display_semantics is None
    assert cells["A13"].value_type is None
    assert cells["A13"].display_semantics == "percentage"
    assert cells["A14"].display_semantics == "currency"
    assert cells["A15"].display_semantics == "date"
    assert cells["A16"].display_semantics == "leading_zero"
    assert (cells["A17"].value_type, cells["A17"].display_semantics) == ("date", "date")
