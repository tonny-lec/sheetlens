from __future__ import annotations

from typing import Any, Literal

from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.xml.constants import MAX_COLUMN, MAX_ROW
from pydantic import BaseModel, Field, computed_field, model_validator

Primitive = str | int | float | bool | None
CellValueType = Literal[
    "string",
    "number",
    "boolean",
    "date",
    "time",
    "datetime",
    "duration",
    "error",
]
CellDisplaySemantics = Literal[
    "percentage",
    "currency",
    "date",
    "time",
    "datetime",
    "duration",
    "leading_zero",
    "error",
]


def _normalized_range_boundaries(reference: str) -> tuple[int, int, int, int]:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(reference)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"range が不正です: {reference}") from exc
    normalized = (
        min_col or 1,
        min_row or 1,
        max_col or MAX_COLUMN,
        max_row or MAX_ROW,
    )
    if not (
        1 <= normalized[0] <= normalized[2] <= MAX_COLUMN
        and 1 <= normalized[1] <= normalized[3] <= MAX_ROW
    ):
        raise ValueError(f"range が Excel の境界外です: {reference}")
    return normalized


def _structural_union(
    content_range: str | None,
    structural_range: str | None,
) -> str | None:
    content = (
        None
        if content_range is None
        else _normalized_range_boundaries(content_range)
    )
    structural = (
        None
        if structural_range is None
        else _normalized_range_boundaries(structural_range)
    )
    if content_range is None:
        return structural_range
    if structural_range is None:
        return content_range
    assert content is not None
    assert structural is not None
    if (
        structural[0] <= content[0]
        and structural[1] <= content[1]
        and content[2] <= structural[2]
        and content[3] <= structural[3]
    ):
        return structural_range
    min_col = min(content[0], structural[0])
    min_row = min(content[1], structural[1])
    max_col = max(content[2], structural[2])
    max_row = max(content[3], structural[3])
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


class Cell(BaseModel):
    ref: str
    value: Primitive = None
    formula: str | None = None
    value_type: CellValueType | None = None
    number_format: str | None = None
    display_semantics: CellDisplaySemantics | None = None


class ValidationRule(BaseModel):
    ranges: list[str]
    type: str
    formula1: str | None = None
    choices: list[str] = Field(default_factory=list)


class ConditionalValue(BaseModel):
    type: str | None = None
    value: str | float | int | None = None
    gte: bool | None = None


class ConditionalColor(BaseModel):
    type: str
    value: str | float | int | bool
    tint: float = 0.0


class ConditionalColorScale(BaseModel):
    conditions: list[ConditionalValue] = Field(default_factory=list)
    colors: list[ConditionalColor] = Field(default_factory=list)


class ConditionalDataBar(BaseModel):
    conditions: list[ConditionalValue] = Field(default_factory=list)
    color: ConditionalColor
    show_value: bool | None = None
    min_length: int | None = None
    max_length: int | None = None


class ConditionalIconSet(BaseModel):
    icon_style: str | None = None
    conditions: list[ConditionalValue] = Field(default_factory=list)
    show_value: bool | None = None
    percent: bool | None = None
    reverse: bool | None = None


class OoxmlNode(BaseModel):
    tag: str
    attributes: dict[str, str] = Field(default_factory=dict)
    text: str | None = None
    children: list[OoxmlNode] = Field(default_factory=list)


class ConditionalFormat(BaseModel):
    range: str
    rule_type: str
    operator: str | None = None
    formulas: list[str] = Field(default_factory=list)
    stop_if_true: bool = False
    color_scale: ConditionalColorScale | None = None
    data_bar: ConditionalDataBar | None = None
    icon_set: ConditionalIconSet | None = None
    dxf: OoxmlNode | None = None

    @model_validator(mode="before")
    @classmethod
    def migrate_legacy_formula(cls, data: Any) -> Any:
        if not isinstance(data, dict):
            return data
        migrated = dict(data)
        legacy_formula = migrated.pop("formula", None)
        if "formulas" not in migrated:
            migrated["formulas"] = [] if legacy_formula is None else [legacy_formula]
        return migrated

    @property
    def formula(self) -> str | None:
        return self.formulas[0] if self.formulas else None

    @formula.setter
    def formula(self, value: str | None) -> None:
        if value is None:
            self.formulas = []
        elif self.formulas:
            self.formulas[0] = value
        else:
            self.formulas = [value]


class VbaModule(BaseModel):
    name: str
    code: str


class ButtonLink(BaseModel):
    sheet: str
    label: str | None = None
    macro: str


class DependencyEdge(BaseModel):
    source: str
    target_workbook: str | None = None
    target_sheet: str | None = None
    target_range: str | None = None
    unresolved: bool = False


class SheetArtifact(BaseModel):
    type: Literal["chart", "image", "shape", "pivot"]
    count: int = Field(gt=0)
    ooxml_parts: list[str] = Field(default_factory=list)

    @model_validator(mode="after")
    def normalize_ooxml_parts(self) -> SheetArtifact:
        self.ooxml_parts = sorted(set(self.ooxml_parts))
        return self


class Sheet(BaseModel):
    name: str
    content_range: str | None = None
    structural_range: str | None = None
    hidden: bool = False
    protected: bool = False
    hidden_cols: list[str] = Field(default_factory=list)
    hidden_rows: list[int] = Field(default_factory=list)
    cells: list[Cell] = Field(default_factory=list)
    merged: list[str] = Field(default_factory=list)
    validations: list[ValidationRule] = Field(default_factory=list)
    conditional_formats: list[ConditionalFormat] = Field(default_factory=list)
    artifacts: list[SheetArtifact] = Field(default_factory=list)

    @model_validator(mode="before")
    @classmethod
    def synchronize_content_range(cls, data: Any) -> Any:
        if not isinstance(data, dict):
            return data
        synchronized = dict(data)
        used_range = synchronized.pop("used_range", None)
        content_range = synchronized.get("content_range")
        if used_range is not None and content_range is not None:
            if used_range != content_range:
                raise ValueError("used_range と content_range が競合しています")
        if used_range is not None:
            synchronized["content_range"] = used_range
        synchronized["structural_range"] = _structural_union(
            synchronized.get("content_range"),
            synchronized.get("structural_range"),
        )
        return synchronized

    def __setattr__(self, name: str, value: Any) -> None:
        if "content_range" not in self.__dict__:
            super().__setattr__(name, value)
            return
        if name == "content_range":
            previous_content = self.content_range
            previous_structural = self.structural_range
            if previous_structural == previous_content:
                structural_range = _structural_union(value, value)
            else:
                structural_range = _structural_union(value, previous_structural)
            super().__setattr__(name, value)
            super().__setattr__("structural_range", structural_range)
            return
        if name == "structural_range":
            value = _structural_union(self.content_range, value)
        super().__setattr__(name, value)

    @computed_field
    @property
    def used_range(self) -> str | None:
        return self.content_range

    @used_range.setter
    def used_range(self, value: str | None) -> None:
        self.content_range = value


class Workbook(BaseModel):
    source_file: str
    sha256: str
    sheets: list[Sheet] = Field(default_factory=list)
    vba_modules: list[VbaModule] = Field(default_factory=list)
    buttons: list[ButtonLink] = Field(default_factory=list)
    defined_names: dict[str, str] = Field(default_factory=dict)
    external_refs: list[str] = Field(default_factory=list)
    extraction_gaps: list[str] = Field(default_factory=list)
