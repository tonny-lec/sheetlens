import hashlib
import re
from datetime import date, datetime, time, timedelta
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell as OpenpyxlCell
from openpyxl.styles import numbers
from openpyxl.utils.cell import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.xml.constants import MAX_COLUMN, MAX_ROW

from sheetlens.model import ir
from sheetlens.reader.artifacts import extract_sheet_artifacts
from sheetlens.reader.buttons import extract_buttons
from sheetlens.reader.features import read_conditional_formats, read_validations
from sheetlens.reader.vba import extract_vba


def _coerce(value: object) -> ir.Primitive:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)  # datetime 等は文字列化


def _formula_text(raw: object) -> str | None:
    if isinstance(raw, str):
        return raw
    text = getattr(raw, "text", None)
    return text if isinstance(text, str) else None


def _datetime_format_semantics(number_format: str) -> ir.CellDisplaySemantics | None:
    normalized_format = number_format.lower()
    if numbers.is_timedelta_format(normalized_format):
        return "duration"
    semantic = numbers.is_datetime(normalized_format)
    if semantic in ("date", "time", "datetime"):
        return semantic
    return None


def _datetime_value_type(value: object, number_format: str) -> ir.CellValueType | None:
    if not isinstance(value, (datetime, date, time, timedelta)):
        return None
    semantic = _datetime_format_semantics(number_format)
    if semantic is not None:
        return semantic
    if isinstance(value, timedelta):
        return "duration"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    return "time"


def _active_format_text(number_format: str) -> str:
    active: list[str] = []
    in_quote = False
    index = 0
    while index < len(number_format):
        char = number_format[index]
        if char == '"':
            in_quote = not in_quote
            index += 1
            continue
        if in_quote:
            index += 1
            continue
        if char in ("\\", "_", "*"):
            index += 2
            continue
        active.append(char)
        index += 1
    return "".join(active)


def _has_currency_format(number_format: str) -> bool:
    if re.search(r"\[\$(?!-)[^\]]+\]", number_format):
        return True
    without_brackets = re.sub(r"\[[^\]]*\]", "", number_format)
    return any(symbol in without_brackets for symbol in ("$", "¥", "￥", "€", "£"))


def _has_leading_zero_format(number_format: str) -> bool:
    active = _active_format_text(number_format).split(";", 1)[0]
    active = re.sub(r"\[[^\]]*\]", "", active)
    if "/" in active or re.search(r"[Ee][+-]?0", active):
        return False
    integer_part = active.split(".", 1)[0]
    return re.search(r"0{2,}", integer_part) is not None


def _value_type(
    cell: OpenpyxlCell,
    value: object,
    number_format: str,
) -> ir.CellValueType | None:
    if value is None:
        return None
    if cell.data_type == "e":
        return "error"
    datetime_value_type = _datetime_value_type(value, number_format)
    if datetime_value_type is not None:
        return datetime_value_type
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, str):
        return "string"
    return None


def _display_semantics(
    value_type: ir.CellValueType | None,
    value: object,
    number_format: str,
) -> ir.CellDisplaySemantics | None:
    if value_type in ("error", "date", "time", "datetime", "duration"):
        return value_type
    if value_type == "string" and isinstance(value, str) and re.fullmatch(r"0\d+", value):
        return "leading_zero"
    if value_type not in ("number", None):
        return None
    datetime_semantics = _datetime_format_semantics(number_format)
    if datetime_semantics is not None:
        return datetime_semantics
    active_format = _active_format_text(number_format)
    if "%" in active_format:
        return "percentage"
    if _has_currency_format(number_format):
        return "currency"
    if _has_leading_zero_format(number_format):
        return "leading_zero"
    return None


def _read_cell(formula_cell: OpenpyxlCell, value_cell: OpenpyxlCell) -> ir.Cell:
    is_formula = formula_cell.data_type == "f"
    raw_value = value_cell.value if is_formula else formula_cell.value
    metadata_cell = value_cell if is_formula else formula_cell
    number_format = formula_cell.number_format
    value_type = _value_type(metadata_cell, raw_value, number_format)
    return ir.Cell(
        ref=formula_cell.coordinate,
        value=_coerce(raw_value),
        formula=_formula_text(formula_cell.value) if is_formula else None,
        value_type=value_type,
        number_format=number_format,
        display_semantics=_display_semantics(value_type, raw_value, number_format),
    )


def _range_envelope(
    references: list[str],
    *,
    sheet_name: str,
    gaps: list[str],
) -> str | None:
    bounds: list[tuple[int, int, int, int]] = []
    for reference in references:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(reference)
        except (TypeError, ValueError):
            gaps.append(f"{sheet_name}!{reference}: 構造範囲を解釈できませんでした")
            continue
        normalized = (
            min_col or 1,
            min_row or 1,
            max_col or MAX_COLUMN,
            max_row or MAX_ROW,
        )
        if not (
            1 <= normalized[0] <= normalized[2] <= MAX_COLUMN
            and 1 <= normalized[1] <= normalized[3] <= MAX_ROW
        ):
            gaps.append(f"{sheet_name}!{reference}: 構造範囲が Excel の境界外です")
            continue
        bounds.append(
            normalized
        )
    if not bounds:
        return None
    min_col = min(bound[0] for bound in bounds)
    min_row = min(bound[1] for bound in bounds)
    max_col = max(bound[2] for bound in bounds)
    max_row = max(bound[3] for bound in bounds)
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _raw_structural_ranges(ws, gaps: list[str]) -> list[str]:
    references = [str(cell_range) for cell_range in ws.merged_cells.ranges]
    try:
        for validation in ws.data_validations.dataValidation:
            references.extend(str(cell_range) for cell_range in validation.sqref.ranges)
    except Exception as exc:  # noqa: BLE001 — 構造情報の欠落を gap に残す
        gaps.append(f"{ws.title}: 入力規則の構造範囲抽出に失敗 ({exc})")
    try:
        for conditional_format in ws.conditional_formatting:
            references.extend(
                str(cell_range) for cell_range in conditional_format.sqref.ranges
            )
    except Exception as exc:  # noqa: BLE001 — 構造情報の欠落を gap に残す
        gaps.append(f"{ws.title}: 条件付き書式の構造範囲抽出に失敗 ({exc})")
    return references


def _hidden_columns(ws) -> list[str]:
    hidden: set[int] = set()
    for key, dimension in ws.column_dimensions.items():
        if not dimension.hidden:
            continue
        start = dimension.min or column_index_from_string(key)
        end = dimension.max or start
        hidden.update(range(start, end + 1))
    return [get_column_letter(index) for index in sorted(hidden)]


def _materialized_value_cells(ws) -> list[OpenpyxlCell]:
    # iter_rows() は書式だけの最遠セルまで矩形展開するため、実体化済みセルだけを見る。
    return sorted(
        (
            cell
            for cell in ws._cells.values()  # noqa: SLF001 — openpyxl に公開 iterator がない
            if isinstance(cell, OpenpyxlCell) and cell.value is not None
        ),
        key=lambda cell: (cell.row, cell.column),
    )


def read_workbook(path: Path) -> ir.Workbook:
    data = path.read_bytes()
    keep_vba = path.suffix.lower() in (".xlsm", ".xltm")
    wb_f = openpyxl.load_workbook(path, data_only=False, keep_vba=keep_vba)
    wb_v = openpyxl.load_workbook(path, data_only=True)
    gaps: list[str] = []
    artifacts_by_sheet, artifact_gaps = extract_sheet_artifacts(path)
    gaps.extend(artifact_gaps)
    sheets: list[ir.Sheet] = []
    for ws_f in wb_f.worksheets:
        ws_v = wb_v[ws_f.title]
        cells: list[ir.Cell] = []
        for c in _materialized_value_cells(ws_f):
            if c.data_type == "f":
                raw = c.value
                formula = _formula_text(raw)
                if formula is None:
                    gaps.append(
                        f"{ws_f.title}!{c.coordinate}: 未対応の数式型 "
                        f"{type(raw).__name__} のため数式を抽出できませんでした"
                    )
            cells.append(_read_cell(c, ws_v[c.coordinate]))
        try:
            validations = read_validations(
                ws_f,
                wb_v,
                extraction_gaps=gaps,
            )
        except Exception as e:  # noqa: BLE001 — 欠落は gap として記録して継続
            validations = []
            gaps.append(f"{ws_f.title}: 入力規則の抽出に失敗 ({e})")
        try:
            cformats = read_conditional_formats(ws_f, extraction_gaps=gaps)
        except Exception as e:  # noqa: BLE001
            cformats = []
            gaps.append(f"{ws_f.title}: 条件付き書式の抽出に失敗 ({e})")
        content_range = _range_envelope(
            [cell.ref for cell in cells],
            sheet_name=ws_f.title,
            gaps=gaps,
        )
        structural_range = _range_envelope(
            [cell.ref for cell in cells] + _raw_structural_ranges(ws_f, gaps),
            sheet_name=ws_f.title,
            gaps=gaps,
        )
        sheets.append(
            ir.Sheet(
                name=ws_f.title,
                used_range=content_range,
                content_range=content_range,
                structural_range=structural_range,
                hidden=ws_f.sheet_state != "visible",
                protected=bool(ws_f.protection.sheet),
                hidden_cols=_hidden_columns(ws_f),
                hidden_rows=sorted(k for k, v in ws_f.row_dimensions.items() if v.hidden),
                cells=cells,
                merged=[str(r) for r in ws_f.merged_cells.ranges],
                validations=validations,
                conditional_formats=cformats,
                artifacts=artifacts_by_sheet.get(ws_f.title, []),
            )
        )
    defined = {}
    for name, defn in wb_f.defined_names.items():
        defined[name] = defn.attr_text or ""
    vba_modules: list[ir.VbaModule] = []
    buttons: list[ir.ButtonLink] = []
    try:
        vba_modules = extract_vba(path)
    except Exception as e:  # noqa: BLE001
        gaps.append(f"VBA の抽出に失敗 ({e})")
    try:
        buttons = extract_buttons(path, extraction_gaps=gaps)
    except Exception as e:  # noqa: BLE001
        gaps.append(f"ボタン↔マクロ対応の抽出に失敗 ({e})")
    return ir.Workbook(
        source_file=path.name,
        sha256=hashlib.sha256(data).hexdigest(),
        sheets=sheets,
        vba_modules=vba_modules,
        buttons=buttons,
        defined_names=defined,
        extraction_gaps=gaps,
    )
