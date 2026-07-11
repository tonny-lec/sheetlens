from dataclasses import dataclass

from openpyxl.utils.cell import SHEETRANGE_RE, range_boundaries
from openpyxl.xml.constants import MAX_COLUMN, MAX_ROW

from sheetlens.model import ir


@dataclass(frozen=True)
class _RangeTarget:
    sheet: str
    min_col: int
    min_row: int
    max_col: int
    max_row: int


@dataclass(frozen=True)
class _ListResolution:
    choices: list[str]
    reason: str | None = None


def _formula_source(formula: str) -> str:
    source = formula.strip()
    if source.startswith("="):
        source = source[1:].strip()
    return source


def _unsupported_function_reason(source: str) -> str | None:
    upper = source.lstrip().upper()
    if upper.startswith("INDIRECT("):
        return "unsupported_indirect"
    if upper.startswith("OFFSET("):
        return "unsupported_offset"
    return None


def _is_defined_name_candidate(source: str) -> bool:
    if not source or not (source[0].isalpha() or source[0] in "_\\"):
        return False
    return all(char.isalnum() or char in "._\\" for char in source[1:])


def _format_validation_gap(
    sheet: str,
    ranges: list[str],
    formula1: str,
    reason: str,
) -> str:
    sorted_ranges = ", ".join(sorted(ranges))
    return (
        f"{sheet}: 入力規則 {sorted_ranges} の選択肢を解決できません "
        f"(formula1={formula1!r}; reason={reason})"
    )


def _parse_static_range(source: str, default_sheet: str | None) -> _RangeTarget | None:
    try:
        if "!" in source:
            match = SHEETRANGE_RE.fullmatch(source)
            if match is None:
                return None
            sheet = match.group("quoted") or match.group("notquoted")
            sheet = sheet.replace("''", "'")
            bounds = range_boundaries(match.group("cells"))
        else:
            if default_sheet is None:
                return None
            sheet = default_sheet
            bounds = range_boundaries(source)
    except (AttributeError, TypeError, ValueError):
        return None

    if any(value is None for value in bounds):
        return None
    min_col, min_row, max_col, max_row = bounds
    if not (
        1 <= min_col <= max_col <= MAX_COLUMN
        and 1 <= min_row <= max_row <= MAX_ROW
    ):
        return None
    return _RangeTarget(sheet, min_col, min_row, max_col, max_row)


def _read_range(wb_v, target: _RangeTarget) -> _ListResolution:
    if any(marker in target.sheet for marker in (":", "[", "]")):
        return _ListResolution([], "unsupported_reference")
    if target.sheet not in wb_v.sheetnames:
        return _ListResolution([], "sheet_not_found")
    ws = wb_v[target.sheet]
    choices = [
        str(cell.value)
        for row in ws.iter_rows(
            min_col=target.min_col,
            min_row=target.min_row,
            max_col=target.max_col,
            max_row=target.max_row,
        )
        for cell in row
        if cell.value is not None
    ]
    return _ListResolution(choices)


def _find_defined_name(mapping, name: str):
    matches = [
        definition
        for key, definition in mapping.items()
        if key.casefold() == name.casefold()
    ]
    if len(matches) > 1:
        return None, "ambiguous_name"
    return (matches[0], None) if matches else (None, None)


def _resolve_definition(wb_v, definition, *, default_sheet: str | None) -> _ListResolution:
    source = _formula_source(definition.attr_text or "")
    function_reason = _unsupported_function_reason(source)
    if function_reason is not None:
        return _ListResolution([], function_reason)
    target = _parse_static_range(source, default_sheet)
    if target is not None:
        return _read_range(wb_v, target)
    if "," in source:
        return _ListResolution([], "unsupported_reference")
    reason = (
        "invalid_range"
        if "!" in source or "#REF!" in source
        else "unsupported_reference"
    )
    return _ListResolution([], reason)


def _resolve_list(wb_v, current_sheet: str, formula: str) -> _ListResolution:
    source = _formula_source(formula)
    if source.startswith("="):
        return _ListResolution([], "unsupported_reference")

    target = _parse_static_range(source, current_sheet)
    if target is not None:
        return _read_range(wb_v, target)

    function_reason = _unsupported_function_reason(source)
    if function_reason is not None:
        return _ListResolution([], function_reason)
    if "," in source:
        return _ListResolution([], "unsupported_reference")
    if "!" in source or "#REF!" in source:
        return _ListResolution([], "invalid_range")
    if not _is_defined_name_candidate(source):
        return _ListResolution([], "unsupported_reference")

    if current_sheet not in wb_v.sheetnames:
        return _ListResolution([], "sheet_not_found")

    local_definition, reason = _find_defined_name(
        wb_v[current_sheet].defined_names,
        source,
    )
    if reason is not None:
        return _ListResolution([], reason)
    if local_definition is not None:
        return _resolve_definition(
            wb_v,
            local_definition,
            default_sheet=current_sheet,
        )

    workbook_definition, reason = _find_defined_name(wb_v.defined_names, source)
    if reason is not None:
        return _ListResolution([], reason)
    if workbook_definition is None:
        return _ListResolution([], "name_not_found")
    return _resolve_definition(wb_v, workbook_definition, default_sheet=None)


def read_validations(
    ws_f,
    wb_v,
    *,
    extraction_gaps: list[str] | None = None,
) -> list[ir.ValidationRule]:
    gap_sink = [] if extraction_gaps is None else extraction_gaps
    rules: list[ir.ValidationRule] = []
    for dv in ws_f.data_validations.dataValidation:
        f1 = dv.formula1
        ranges = [str(r) for r in dv.sqref.ranges]
        choices: list[str] = []
        if dv.type == "list" and f1:
            if f1.startswith('"'):
                choices = [s.strip() for s in f1.strip('"').split(",")]
            else:
                resolution = _resolve_list(wb_v, ws_f.title, f1)
                choices = resolution.choices
                if resolution.reason is not None:
                    gap_sink.append(
                        _format_validation_gap(
                            ws_f.title,
                            ranges,
                            f1,
                            resolution.reason,
                        )
                    )
        rules.append(
            ir.ValidationRule(
                ranges=ranges,
                type=dv.type or "unknown",
                formula1=f1,
                choices=choices,
            )
        )
    return rules


def _normalize_xml_node(element) -> ir.OoxmlNode:
    return ir.OoxmlNode(
        tag=str(element.tag),
        attributes={str(key): str(value) for key, value in element.attrib.items()},
        text=None if element.text is None else str(element.text),
        children=[_normalize_xml_node(child) for child in element],
    )


def _format_conditional_format_gap(
    sheet: str,
    target_range: str,
    rule_type: str,
    reason: str,
) -> str:
    return (
        f"{sheet}: 条件付き書式 {target_range} を完全に抽出できません "
        f"(type={rule_type}; reason={reason})"
    )


def _normalize_conditional_value(value) -> ir.ConditionalValue:
    raw_value = getattr(value, "val")
    if isinstance(raw_value, bool) or (
        raw_value is not None and not isinstance(raw_value, (str, int, float))
    ):
        raise ValueError("invalid conditional value")
    return ir.ConditionalValue(
        type=getattr(value, "type"),
        value=raw_value,
        gte=getattr(value, "gte", None),
    )


def _normalize_conditional_color(color) -> ir.ConditionalColor:
    color_type = getattr(color, "type")
    value = getattr(color, "value")
    if color_type not in {"rgb", "theme", "indexed", "auto"}:
        raise ValueError("invalid color type")
    if not isinstance(value, (str, int, float, bool)):
        raise ValueError("invalid color value")
    return ir.ConditionalColor(
        type=color_type,
        value=value,
        tint=float(getattr(color, "tint", 0.0)),
    )


def _normalize_color_scale(payload) -> ir.ConditionalColorScale:
    values = list(getattr(payload, "cfvo"))
    colors = list(getattr(payload, "color"))
    if not values or len(values) != len(colors):
        raise ValueError("conditions and colors must be non-empty and aligned")
    return ir.ConditionalColorScale(
        conditions=[_normalize_conditional_value(value) for value in values],
        colors=[_normalize_conditional_color(color) for color in colors],
    )


def _normalize_data_bar(payload) -> ir.ConditionalDataBar:
    values = list(getattr(payload, "cfvo"))
    if len(values) != 2:
        raise ValueError("data bar requires start and end conditions")
    color = getattr(payload, "color")
    if color is None:
        raise ValueError("data bar requires a color")
    return ir.ConditionalDataBar(
        conditions=[_normalize_conditional_value(value) for value in values],
        color=_normalize_conditional_color(color),
        show_value=getattr(payload, "showValue", None),
        min_length=getattr(payload, "minLength", None),
        max_length=getattr(payload, "maxLength", None),
    )


def _normalize_icon_set(payload) -> ir.ConditionalIconSet:
    values = list(getattr(payload, "cfvo"))
    icon_style = getattr(payload, "iconSet")
    if not values or not icon_style:
        raise ValueError("icon set requires a style and conditions")
    return ir.ConditionalIconSet(
        icon_style=icon_style,
        conditions=[_normalize_conditional_value(value) for value in values],
        show_value=getattr(payload, "showValue", None),
        percent=getattr(payload, "percent", None),
        reverse=getattr(payload, "reverse", None),
    )


def _normalize_visual_payload(rule, rule_type: str):
    if rule_type == "colorScale":
        payload = getattr(rule, "colorScale", None)
        if payload is None:
            return {}, "missing_color_scale"
        try:
            return {"color_scale": _normalize_color_scale(payload)}, None
        except (AttributeError, TypeError, ValueError):
            return {}, "invalid_color_scale"
    if rule_type == "dataBar":
        payload = getattr(rule, "dataBar", None)
        if payload is None:
            return {}, "missing_data_bar"
        try:
            return {"data_bar": _normalize_data_bar(payload)}, None
        except (AttributeError, TypeError, ValueError):
            return {}, "invalid_data_bar"
    if rule_type == "iconSet":
        payload = getattr(rule, "iconSet", None)
        if payload is None:
            return {}, "missing_icon_set"
        try:
            return {"icon_set": _normalize_icon_set(payload)}, None
        except (AttributeError, TypeError, ValueError):
            return {}, "invalid_icon_set"
    if rule_type not in {"cellIs", "expression"}:
        return {}, "unsupported_type"
    return {}, None


def read_conditional_formats(
    ws_f,
    *,
    extraction_gaps: list[str] | None = None,
) -> list[ir.ConditionalFormat]:
    gap_sink = [] if extraction_gaps is None else extraction_gaps
    out: list[ir.ConditionalFormat] = []
    for fmt in ws_f.conditional_formatting:
        for rule in fmt.rules:
            target_range = "unknown"
            rule_type = "unknown"
            formulas: list[str] = []
            operator = None
            stop_if_true = False
            dxf = None
            visual_payload = {}
            reason: str | None = None

            try:
                target_range = str(fmt.sqref)
            except Exception:  # noqa: BLE001 — retain a minimal rule and continue
                reason = "extraction_error"

            type_available = True
            try:
                rule_type = getattr(rule, "type", None) or "unknown"
            except Exception:  # noqa: BLE001 — retain a minimal rule and continue
                type_available = False
                reason = "extraction_error"

            try:
                formulas = [str(formula) for formula in (getattr(rule, "formula", None) or [])]
            except Exception:  # noqa: BLE001 — retain fields extracted in other stages
                reason = "extraction_error"

            try:
                raw_dxf = getattr(rule, "dxf", None)
                if raw_dxf is not None:
                    dxf = _normalize_xml_node(raw_dxf.to_tree())
            except Exception:  # noqa: BLE001 — preserve the rest of the rule
                if reason is None:
                    reason = "invalid_dxf"

            if type_available:
                try:
                    visual_payload, payload_reason = _normalize_visual_payload(
                        rule,
                        rule_type,
                    )
                    if payload_reason is not None:
                        if reason in {None, "invalid_dxf"}:
                            reason = payload_reason
                except Exception:  # noqa: BLE001 — preserve earlier extraction stages
                    reason = "extraction_error"

            try:
                operator = getattr(rule, "operator", None)
            except Exception:  # noqa: BLE001 — preserve earlier extraction stages
                reason = "extraction_error"

            try:
                stop_if_true = bool(getattr(rule, "stopIfTrue", False))
            except Exception:  # noqa: BLE001 — preserve earlier extraction stages
                reason = "extraction_error"

            try:
                out.append(
                    ir.ConditionalFormat(
                        range=target_range,
                        rule_type=rule_type,
                        operator=operator,
                        formulas=formulas,
                        stop_if_true=stop_if_true,
                        dxf=dxf,
                        **visual_payload,
                    )
                )
            except Exception:  # noqa: BLE001 — isolate each rule and keep later rules
                reason = "extraction_error"
                out.append(
                    ir.ConditionalFormat(
                        range=target_range,
                        rule_type=rule_type,
                    )
                )
            if reason is not None:
                gap_sink.append(
                    _format_conditional_format_gap(
                        ws_f.title,
                        target_range,
                        rule_type,
                        reason,
                    )
                )
    return out
