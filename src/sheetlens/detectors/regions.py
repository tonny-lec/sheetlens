from collections import defaultdict

from openpyxl.utils import coordinate_to_tuple, get_column_letter, range_boundaries
from pydantic import BaseModel

from sheetlens.detectors.util import runs
from sheetlens.model import ir


class Region(BaseModel):
    range: str
    kind: str  # "table" | "block"


def detect_regions(sheet: ir.Sheet) -> list[Region]:
    rows: dict[int, list[tuple[int, ir.Cell]]] = defaultdict(list)
    for cell in sheet.cells:
        r, c = coordinate_to_tuple(cell.ref)
        rows[r].append((c, cell))
    if not rows:
        return []
    regions: list[Region] = []
    for row_start, row_end in runs(sorted(rows)):
        occupied_cols = sorted(
            {column for row in range(row_start, row_end + 1) for column, _ in rows[row]}
        )
        first_row = [cell for _, cell in rows[row_start]]
        has_table_header = (
            row_end - row_start + 1 >= 3
            and len(first_row) >= 2
            and all(
                isinstance(cell.value, str) and cell.formula is None
                for cell in first_row
            )
        )
        column_bands = runs(occupied_cols)
        if has_table_header:
            data_cols = sorted(
                {
                    column
                    for row in range(row_start + 1, row_end + 1)
                    for column, _ in rows[row]
                }
            )
            header_only_cols = sorted(set(occupied_cols) - set(data_cols))
            column_bands = sorted([*runs(data_cols), *runs(header_only_cols)])

        for col_start, col_end in column_bands:
            occupied_rows = sorted(
                row
                for row in range(row_start, row_end + 1)
                if any(col_start <= column <= col_end for column, _ in rows[row])
            )
            for start, end in runs(occupied_rows):
                cols = [
                    column
                    for row in range(start, end + 1)
                    for column, _ in rows[row]
                    if col_start <= column <= col_end
                ]
                min_col, max_col = min(cols), max(cols)
                rng = (
                    f"{get_column_letter(min_col)}{start}:"
                    f"{get_column_letter(max_col)}{end}"
                )
                head = [
                    cell
                    for column, cell in sorted(rows[start], key=lambda item: item[0])
                    if min_col <= column <= max_col
                ]
                is_table = (
                    end - start + 1 >= 3
                    and len(head) >= 2
                    and all(
                        isinstance(cell.value, str) and cell.formula is None
                        for cell in head
                    )
                )
                regions.append(Region(range=rng, kind="table" if is_table else "block"))
    return regions


def input_ranges(sheet: ir.Sheet, region: Region) -> list[str]:
    """Return deterministic manual-input portions of a detected region."""
    min_col, min_row, max_col, max_row = range_boundaries(region.range)
    cells: list[tuple[int, int, ir.Cell]] = []
    for cell in sheet.cells:
        row, column = coordinate_to_tuple(cell.ref)
        if min_row <= row <= max_row and min_col <= column <= max_col:
            cells.append((row, column, cell))

    if not any(cell.formula is not None for _, _, cell in cells):
        return [region.range]

    data_min_row = min_row + 1 if region.kind == "table" else min_row
    pure_manual_cols: list[int] = []
    mixed_manual_runs: dict[tuple[int, int], list[int]] = defaultdict(list)
    for column in range(min_col, max_col + 1):
        column_cells = [
            (row, cell)
            for row, cell_column, cell in cells
            if cell_column == column and row >= data_min_row
        ]
        formula_rows = {
            row for row, cell in column_cells if cell.formula is not None
        }
        manual_rows = sorted(
            row
            for row, cell in column_cells
            if cell.formula is None and cell.value is not None
        )
        if not formula_rows:
            if manual_rows:
                pure_manual_cols.append(column)
            continue
        for row_run in runs(manual_rows):
            mixed_manual_runs[row_run].append(column)

    projected: list[tuple[int, int, int, int]] = [
        (start_col, min_row, end_col, max_row)
        for start_col, end_col in runs(pure_manual_cols)
    ]
    projected.extend(
        (start_col, start_row, end_col, end_row)
        for (start_row, end_row), columns in mixed_manual_runs.items()
        for start_col, end_col in runs(columns)
    )
    projected.sort()
    return [
        f"{get_column_letter(start_col)}{start_row}:"
        f"{get_column_letter(end_col)}{end_row}"
        for start_col, start_row, end_col, end_row in projected
    ]
