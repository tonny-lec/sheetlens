from collections.abc import Iterable
import re

from openpyxl.utils import get_column_letter, range_boundaries

from sheetlens.annotations.schema import AnnotationTarget, SheetAnnotations, split_ranges
from sheetlens.detectors.formula_patterns import FormulaPattern
from sheetlens.detectors.questions import Question
from sheetlens.detectors.regions import Region
from sheetlens.model import ir

MAX_GRID_ROWS = 40
MAX_GRID_COLS = 15


def _fmt_target(t: AnnotationTarget) -> str:
    if t.kind == "input_source":
        value = getattr(t, "value", None)
        parts = [f"入力元: {value}"] if value else []
        by = getattr(t, "by", None)
        when = getattr(t, "when", None)
        if by:
            parts.append(f"入力者: {by}")
        if when:
            parts.append(f"タイミング: {when}")
        return " / ".join(parts) or (getattr(t, "note", None) or "")
    if t.kind == "trigger_timing":
        when = getattr(t, "when", None)
        note = getattr(t, "note", None)
        return " / ".join(
            part for part in (f"タイミング: {when}" if when else "", note or "") if part
        )
    values = getattr(t, "values", None)
    if t.kind == "dropdown_semantics" and values:
        return "選択肢の意味: " + "、".join(f"「{k}」={v}" for k, v in values.items())
    return getattr(t, "note", None) or getattr(t, "value", None) or ""


def _ann_lines(ann: SheetAnnotations | None, rng: str) -> list[str]:
    if not ann:
        return []
    lines = [
        f"> 💬 業務上の意味: {_fmt_target(t)}"
        for t in ann.targets
        if getattr(t, "range", None) and rng in split_ranges(t.range)
    ]
    return lines + [""] if lines else []


def _cell_text(text: str) -> str:
    return text.replace("\r\n", " ").replace("\n", " ").replace("\r", " ").replace("|", "\\|")


def _inline_code(text: str) -> str:
    text = text.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    longest_run = max((len(run) for run in re.findall(r"`+", text)), default=0)
    delimiter = "`" * (longest_run + 1)
    needs_padding = (
        text.startswith(("`", " ")) or text.endswith(("`", " "))
    ) and not text.isspace()
    padding = " " if needs_padding else ""
    return f"{delimiter}{padding}{text}{padding}{delimiter}"


def _display_semantics_lines(sheet: ir.Sheet) -> list[str]:
    semantic_cells = [cell for cell in sheet.cells if cell.display_semantics is not None]
    if not semantic_cells:
        return []
    lines = ["## セル表示情報", ""]
    for cell in semantic_cells:
        value_type = cell.value_type or "unknown"
        number_format = _inline_code(cell.number_format or "General")
        lines.append(
            f"- {cell.ref}: {cell.display_semantics} / value_type={value_type} / "
            f"number_format={number_format}"
        )
    return lines + [""]


def _conditional_value_text(value: ir.ConditionalValue) -> str:
    text = value.type or "unknown"
    if value.value is not None:
        text += f"={_conditional_primitive_text(value.value)}"
    if value.gte is not None:
        text += f" (gte={_conditional_primitive_text(value.gte)})"
    return text


def _conditional_color_text(color: ir.ConditionalColor) -> str:
    text = f"{color.type}={_conditional_primitive_text(color.value)}"
    if color.tint:
        text += f" (tint={color.tint})"
    return text


def _conditional_primitive_text(value: ir.Primitive) -> str:
    if isinstance(value, bool):
        return str(value).lower()
    return str(value)


def _conditional_payload_text(cf: ir.ConditionalFormat) -> str | None:
    if cf.color_scale:
        conditions = ", ".join(_conditional_value_text(v) for v in cf.color_scale.conditions)
        colors = ", ".join(_conditional_color_text(c) for c in cf.color_scale.colors)
        return f"colorScale(conditions=[{conditions}], colors=[{colors}])"
    if cf.data_bar:
        conditions = ", ".join(_conditional_value_text(v) for v in cf.data_bar.conditions)
        color = _conditional_color_text(cf.data_bar.color)
        return (
            f"dataBar(conditions=[{conditions}], color={color}, "
            f"showValue={_conditional_primitive_text(cf.data_bar.show_value)}, "
            f"length={cf.data_bar.min_length}-{cf.data_bar.max_length})"
        )
    if cf.icon_set:
        conditions = ", ".join(_conditional_value_text(v) for v in cf.icon_set.conditions)
        return (
            f"iconSet(style={cf.icon_set.icon_style}, conditions=[{conditions}], "
            f"showValue={_conditional_primitive_text(cf.icon_set.show_value)}, "
            f"percent={_conditional_primitive_text(cf.icon_set.percent)}, "
            f"reverse={_conditional_primitive_text(cf.icon_set.reverse)})"
        )
    return None


def _conditional_format_text(cf: ir.ConditionalFormat) -> str:
    formulas = ", ".join(cf.formulas)
    common = " ".join(part for part in (cf.operator, formulas) if part)
    payload = _conditional_payload_text(cf)
    return " / ".join(part for part in (common, payload) if part) or cf.rule_type


def _grid(sheet: ir.Sheet) -> list[str]:
    content_range = sheet.content_range or sheet.used_range
    if not sheet.cells or not content_range:
        return ["（空シート）", ""]
    cellmap = {c.ref: c for c in sheet.cells}
    anchors: dict[str, str] = {}
    covered: set[tuple[int, int]] = set()
    for m in sheet.merged:
        min_c, min_r, max_c, max_r = range_boundaries(m)
        anchors[f"{get_column_letter(min_c)}{min_r}"] = m
        covered.update(
            (r, c)
            for r in range(min_r, max_r + 1)
            for c in range(min_c, max_c + 1)
            if (r, c) != (min_r, min_c)
        )
    min_c, min_r, max_c, max_r = range_boundaries(content_range)
    trunc = max_r - min_r + 1 > MAX_GRID_ROWS or max_c - min_c + 1 > MAX_GRID_COLS
    max_r = min(max_r, min_r + MAX_GRID_ROWS - 1)
    max_c = min(max_c, min_c + MAX_GRID_COLS - 1)
    cols = [get_column_letter(c) for c in range(min_c, max_c + 1)]
    lines = ["| 行 | " + " | ".join(cols) + " |", "|" + "---|" * (len(cols) + 1)]
    for r in range(min_r, max_r + 1):
        row: list[str] = []
        for c in range(min_c, max_c + 1):
            ref = f"{get_column_letter(c)}{r}"
            if (r, c) in covered:
                row.append("←")
                continue
            cell = cellmap.get(ref)
            text = "" if cell is None else str(cell.value if cell.value is not None else cell.formula or "")
            if ref in anchors:
                text = f"[{anchors[ref]} 結合] {text}".strip()
            row.append(_cell_text(text))
        lines.append(f"| {r} | " + " | ".join(row) + " |")
    if trunc:
        lines.append("")
        lines.append(f"（{MAX_GRID_ROWS} 行 × {MAX_GRID_COLS} 列で打ち切り。以降は raw.json を参照）")
    return lines + [""]


def render_sheet_md(
    sheet: ir.Sheet,
    patterns: list[FormulaPattern],
    regions: list[Region],
    questions: list[Question],
    buttons: list[ir.ButtonLink],
    ann: SheetAnnotations | None = None,
    answered: Iterable[str] = frozenset(),
) -> str:
    answered = set(answered)
    lines = [f"# シート: {sheet.name}", ""]
    if ann and ann.role:
        stage = f"（工程: {ann.workflow_stage}）" if ann.workflow_stage else ""
        lines += [f"> 💬 業務上の意味: {ann.role}{stage}", ""]
    flags = []
    if sheet.hidden:
        flags.append("非表示シート")
    if sheet.protected:
        flags.append("保護あり")
    if sheet.hidden_cols:
        flags.append(f"非表示列: {', '.join(sheet.hidden_cols)}")
    lines += [
        "## 概要",
        f"- 内容範囲: {sheet.content_range or sheet.used_range or 'なし'} / "
        f"構造範囲: {sheet.structural_range or 'なし'} / 結合セル: {len(sheet.merged)} 箇所 / "
        f"数式セル: {sum(1 for c in sheet.cells if c.formula)} / 入力規則: {len(sheet.validations)}",
    ]
    if flags:
        lines.append(f"- 属性: {' / '.join(flags)}")
    if ann:
        for t in ann.targets:
            if t.kind == "sheet_role":
                lines.append(f"> 💬 業務上の意味: {_fmt_target(t)}")
                lines.append("")
            if t.kind == "hidden_reason":
                target_range = getattr(t, "range", None)
                label = f"（{target_range}）" if target_range else ""
                lines.append(
                    f"> 💬 業務上の意味{label}: "
                    f"{getattr(t, 'note', None) or getattr(t, 'value', None) or ''}"
                )
    lines += ["", "## レイアウトマップ", ""]
    lines += _grid(sheet)
    lines += _display_semantics_lines(sheet)
    if regions:
        lines += ["## 領域", ""]
        for i, region in enumerate(regions, 1):
            kind = "テーブル" if region.kind == "table" else "ブロック"
            lines.append(f"### 領域{i}: {region.range} — {kind}")
            lines += _ann_lines(ann, region.range)
        lines.append("")
    if patterns:
        lines += ["## 数式（正規化済み）", ""]
        for p in patterns:
            lines.append(f"- {', '.join(p.ranges)} = `{p.pattern}`（例: `{p.example}`）")
            for ex in p.exceptions:
                lines.append(f"  - **⚠ 例外: {ex}**（パターンから逸脱。特例または誤りの可能性）")
        lines.append("")
    if sheet.validations:
        lines += ["## 入力規則（プルダウン等）", ""]
        for v in sheet.validations:
            choices = f" 選択肢: {', '.join(v.choices)}" if v.choices else ""
            lines.append(f"- {', '.join(v.ranges)}: {v.type}（{v.formula1 or ''}）{choices}")
            for rng in v.ranges:
                lines += _ann_lines(ann, rng)
        lines.append("")
    if sheet.conditional_formats:
        lines += ["## 条件付き書式", ""]
        for cf in sheet.conditional_formats:
            cond = _conditional_format_text(cf)
            lines.append(f"- {cf.range}: {cf.rule_type} — 条件: {cond}")
            for rng in split_ranges(cf.range):
                lines += _ann_lines(ann, rng)
        lines.append("")
    if buttons:
        lines += ["## VBA との接続", ""]
        for b in buttons:
            label = f"「{b.label}」" if b.label else ""
            lines.append(f"- ボタン{label} → {b.macro}")
            lines += _ann_lines(ann, b.macro)
        lines.append("")
    unanswered = [q for q in questions if q.sheet == sheet.name and q.id not in answered]
    if unanswered:
        lines += ["## 未確認事項", ""]
        for q in unanswered:
            lines.append(f"> ❓ 未確認 ({q.id}): [{q.target}] {q.text}")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def render_readme(
    wb: ir.Workbook,
    deps: dict[str, list[str]],
    questions: list[Question],
    answered: Iterable[str],
    annotations: Iterable[SheetAnnotations] = (),
) -> str:
    answered = set(answered)
    annotations = list(annotations)
    lines = [f"# {wb.source_file} — SheetLens 抽出結果", ""]
    if wb.extraction_gaps:
        lines += [f"**⚠ この抽出には {len(wb.extraction_gaps)} 件の欠落があります:**", ""]
        lines += [f"- {g}" for g in wb.extraction_gaps]
        lines.append("")
    lines += [
        "## 読み方",
        "",
        "- `structure/sheet-*.md`: シートごとの構造ビュー（compile 後は業務注釈も織り込み済み）",
        "- `structure/raw.json`: 省略なしの全抽出データ（機械可読の正）",
        "- `structure/vba/*.bas`: VBA ソース",
        "- `annotations/*.yaml`: 業務上の意味（人間の回答。手で編集してよい唯一の場所）",
        "- `questions.md`: 業務担当者に確認すべき質問リスト",
        "- `question-ids.json`: 安定した質問 ID と旧 ID alias の履歴",
        "",
        "## 質問 ID と注釈",
        "",
        "- 現在の質問 ID は内容から決定的に生成する `q2-<rule>-<16hex>` 形式です。",
        "- 旧形式の `q-NNN` は `question-ids.json` の alias で引き続き解決し、"
        "alias は一度保存した対応先から変更しません。",
        "- SheetLens は質問 ID を含む `annotations/*.yaml` を書き換えません。",
        "- `legacy_source_sha256` は alias を作ったアップグレード前の抽出 snapshot の由来を示す"
        "文脈情報であり、暗号学的な完全性や回答時点を証明するものではありません。",
        "- `question-ids.json` を手で編集しないでください。alias を別の有効な過去/現行 ID へ"
        "意図的に付け替えた場合、SheetLens はその改ざんを検出できません。",
        "",
        "## シート一覧",
        "",
    ]
    vba_annotations = [ann for ann in annotations if ann.sheet == "(VBA)"]
    if vba_annotations:
        sheet_list_index = lines.index("## シート一覧")
        lines[sheet_list_index:sheet_list_index] = [
            "## VBA 注釈",
            "",
            *[
                f"- {target.range}: {_fmt_target(target)}"
                for ann in vba_annotations
                for target in ann.targets
                if getattr(target, "range", None)
            ],
            "",
        ]
    for s in wb.sheets:
        mark = "（非表示）" if s.hidden else ""
        lines.append(
            f"- {s.name}{mark}: 内容範囲 {s.content_range or s.used_range or 'なし'} / "
            f"構造範囲 {s.structural_range or 'なし'}"
        )
    edges = [(a, bs) for a, bs in deps.items() if bs]
    if edges:
        lines += ["", "## シート間依存（参照する側 → される側）", ""]
        lines += [f"- {a} → {', '.join(bs)}" for a, bs in edges]
    unanswered = sum(1 for q in questions if q.id not in answered)
    lines += ["", f"## 未確認事項: {unanswered} 件（`questions.md` 参照）"]
    return "\n".join(lines).rstrip() + "\n"


def render_questions_md(questions: list[Question], answered: Iterable[str]) -> str:
    answered = set(answered)
    lines = [
        "# 業務担当者への質問リスト",
        "",
        "回答は `annotations/<シート名>.yaml` に記録し、`questions_answered` に ID を追加すること。",
        "注釈 target の range 書式: セル範囲は A1 形式（複数はカンマ区切り）。"
        "trigger_timing はマクロ名、hidden_reason は対象（列名など）をそのまま指定。",
        "",
    ]
    for q in questions:
        mark = "x" if q.id in answered else " "
        lines.append(f"- [{mark}] **{q.id}** [{q.sheet} / {q.target}] ({q.category}) {q.text}")
    return "\n".join(lines).rstrip() + "\n"
