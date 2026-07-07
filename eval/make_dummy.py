"""記憶ベースのダミー業務 Excel（見積管理）を生成する。

使い方: uv run python eval/make_dummy.py [出力パス]
中身はユーザーへのヒアリングで現実の業務パターンに近づけていく。
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


def build(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "見積入力"
    ws.merge_cells("A1:H1")
    ws["A1"] = "見積書"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "顧客名"
    ws["A4"] = "見積日"
    ws["A5"] = "区分"
    dv = DataValidation(type="list", formula1="=区分マスタ!$A$2:$A$3")
    dv.add("B5")
    ws.add_data_validation(dv)
    ws["A10"] = "No"
    ws["B10"] = "品名"
    ws["C10"] = "数量"
    ws["D10"] = "単価"
    ws["E10"] = "金額"
    ws["F10"] = "粗利"
    for r in range(11, 31):
        ws[f"A{r}"] = r - 10
        ws[f"D{r}"] = f"=IFERROR(VLOOKUP(B{r},単価マスタ!$A$2:$C$9,3,FALSE),0)"
        ws[f"E{r}"] = f"=C{r}*D{r}"
        ws[f"F{r}"] = f"=E{r}-C{r}*VLOOKUP(B{r},単価マスタ!$A$2:$C$9,2,FALSE)"
    ws["E32"] = "=SUM(E11:E30)"
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(
        "F11:F30", CellIsRule(operator="lessThan", formula=["0"], fill=red)
    )
    master = wb.create_sheet("単価マスタ")
    master.append(["品名", "原価", "単価"])
    for row in (["部品A", 700, 1000], ["部品B", 3500, 5000], ["組立費", 6000, 8000]):
        master.append(row)
    kubun = wb.create_sheet("区分マスタ")
    kubun.append(["区分"])
    kubun.append(["通常"])
    kubun.append(["特急"])
    kubun.sheet_state = "hidden"
    wb.save(path)


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).parent / "見積管理.xlsx"
    build(out)
    print(f"生成しました: {out}")
